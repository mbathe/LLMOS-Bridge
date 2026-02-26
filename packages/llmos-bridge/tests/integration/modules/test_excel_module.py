"""Integration tests — ExcelModule against real temp .xlsx files."""

from __future__ import annotations

from pathlib import Path

import pytest
import openpyxl

from llmos_bridge.modules.excel import ExcelModule


@pytest.fixture
def module() -> ExcelModule:
    return ExcelModule()


@pytest.fixture
def wb_path(tmp_path: Path) -> Path:
    """Create a simple workbook with one sheet and some data."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "Name"
    ws["B1"] = "Value"
    ws["A2"] = "Alice"
    ws["B2"] = 42
    ws["A3"] = "Bob"
    ws["B3"] = 99
    path = tmp_path / "test.xlsx"
    wb.save(str(path))
    return path


@pytest.mark.integration
class TestOpenCloseWorkbook:
    async def test_open_workbook(self, module: ExcelModule, wb_path: Path) -> None:
        result = await module._action_open_workbook({"path": str(wb_path)})
        assert result["sheet_names"] == ["Sheet1"]
        assert result["active_sheet"] == "Sheet1"

    async def test_close_workbook_not_cached(self, module: ExcelModule, wb_path: Path) -> None:
        result = await module._action_close_workbook({"path": str(wb_path)})
        assert result["closed"] is True
        assert result["was_cached"] is False

    async def test_open_then_close(self, module: ExcelModule, wb_path: Path) -> None:
        await module._action_open_workbook({"path": str(wb_path)})
        result = await module._action_close_workbook({"path": str(wb_path)})
        assert result["was_cached"] is True

    async def test_save_workbook(self, module: ExcelModule, wb_path: Path) -> None:
        await module._action_open_workbook({"path": str(wb_path)})
        result = await module._action_save_workbook({"path": str(wb_path)})
        assert result["saved"] is True
        assert wb_path.exists()

    async def test_get_workbook_info(self, module: ExcelModule, wb_path: Path) -> None:
        result = await module._action_get_workbook_info(
            {"path": str(wb_path), "include_named_ranges": False}
        )
        assert result["sheet_names"] == ["Sheet1"]
        assert result["active_sheet"] == "Sheet1"


@pytest.mark.integration
class TestSheetManagement:
    async def test_list_sheets(self, module: ExcelModule, wb_path: Path) -> None:
        result = await module._action_list_sheets({"path": str(wb_path)})
        assert result["sheet_names"] == ["Sheet1"]
        assert result["count"] == 1

    async def test_create_sheet(self, module: ExcelModule, wb_path: Path) -> None:
        result = await module._action_create_sheet({"path": str(wb_path), "name": "NewSheet"})
        assert result["created"] is True
        assert result["sheet_name"] == "NewSheet"
        # Verify persisted
        module2 = ExcelModule()
        info = await module2._action_list_sheets({"path": str(wb_path)})
        assert "NewSheet" in info["sheet_names"]

    async def test_delete_sheet(self, module: ExcelModule, wb_path: Path) -> None:
        # Create then delete
        await module._action_create_sheet({"path": str(wb_path), "name": "ToDelete"})
        result = await module._action_delete_sheet({"path": str(wb_path), "name": "ToDelete"})
        assert result["deleted"] is True

    async def test_delete_nonexistent_sheet_raises(self, module: ExcelModule, wb_path: Path) -> None:
        with pytest.raises(KeyError):
            await module._action_delete_sheet({"path": str(wb_path), "name": "Ghost"})

    async def test_rename_sheet(self, module: ExcelModule, wb_path: Path) -> None:
        result = await module._action_rename_sheet(
            {"path": str(wb_path), "old_name": "Sheet1", "new_name": "Renamed"}
        )
        assert result["renamed"] is True
        assert result["new_name"] == "Renamed"

    async def test_rename_nonexistent_sheet_raises(self, module: ExcelModule, wb_path: Path) -> None:
        with pytest.raises(KeyError):
            await module._action_rename_sheet(
                {"path": str(wb_path), "old_name": "Ghost", "new_name": "X"}
            )

    async def test_copy_sheet(self, module: ExcelModule, wb_path: Path) -> None:
        result = await module._action_copy_sheet(
            {"path": str(wb_path), "source_sheet": "Sheet1", "new_name": "Sheet1_copy"}
        )
        assert result["copied"] is True
        assert result["new_name"] == "Sheet1_copy"

    async def test_get_sheet_info(self, module: ExcelModule, wb_path: Path) -> None:
        result = await module._action_get_sheet_info(
            {"path": str(wb_path), "sheet": "Sheet1", "include_dimensions": True}
        )
        assert result["sheet_name"] == "Sheet1"
        assert result["dimensions"]["max_row"] >= 3
        assert result["dimensions"]["max_column"] >= 2


@pytest.mark.integration
class TestCellOperations:
    async def test_read_cell(self, module: ExcelModule, wb_path: Path) -> None:
        result = await module._action_read_cell(
            {"path": str(wb_path), "sheet": "Sheet1", "cell": "A1"}
        )
        assert result["value"] == "Name"
        assert result["cell"] == "A1"

    async def test_read_numeric_cell(self, module: ExcelModule, wb_path: Path) -> None:
        result = await module._action_read_cell(
            {"path": str(wb_path), "sheet": "Sheet1", "cell": "B2"}
        )
        assert result["value"] == 42

    async def test_read_cell_nonexistent_sheet(self, module: ExcelModule, wb_path: Path) -> None:
        with pytest.raises(KeyError):
            await module._action_read_cell(
                {"path": str(wb_path), "sheet": "Ghost", "cell": "A1"}
            )

    async def test_write_then_read_cell(self, module: ExcelModule, wb_path: Path) -> None:
        await module._action_write_cell(
            {"path": str(wb_path), "sheet": "Sheet1", "cell": "C1", "value": "Written"}
        )
        # Use fresh module to avoid cache
        module2 = ExcelModule()
        result = await module2._action_read_cell(
            {"path": str(wb_path), "sheet": "Sheet1", "cell": "C1"}
        )
        assert result["value"] == "Written"

    async def test_write_cell_nonexistent_sheet(self, module: ExcelModule, wb_path: Path) -> None:
        with pytest.raises(KeyError):
            await module._action_write_cell(
                {"path": str(wb_path), "sheet": "Ghost", "cell": "A1", "value": "x"}
            )


@pytest.mark.integration
class TestRangeOperations:
    async def test_read_range(self, module: ExcelModule, wb_path: Path) -> None:
        result = await module._action_read_range(
            {"path": str(wb_path), "sheet": "Sheet1", "range": "A1:B3"}
        )
        assert result["row_count"] == 3
        assert result["data"][0] == ["Name", "Value"]
        assert result["data"][1] == ["Alice", 42]

    async def test_read_range_as_dict(self, module: ExcelModule, wb_path: Path) -> None:
        result = await module._action_read_range(
            {"path": str(wb_path), "sheet": "Sheet1", "range": "A1:B3",
             "as_dict": True, "include_headers": True}
        )
        assert "data" in result
        assert result["headers"] == ["Name", "Value"]
        assert result["data"][0] == {"Name": "Alice", "Value": 42}

    async def test_read_range_auto(self, module: ExcelModule, wb_path: Path) -> None:
        result = await module._action_read_range(
            {"path": str(wb_path), "sheet": "Sheet1", "range": "auto"}
        )
        assert result["row_count"] >= 3

    async def test_write_range(self, module: ExcelModule, wb_path: Path) -> None:
        data = [["X", "Y"], [1, 2], [3, 4]]
        result = await module._action_write_range(
            {"path": str(wb_path), "sheet": "Sheet1", "start_cell": "D1", "data": data}
        )
        assert result["written"] is True
        assert result["rows"] == 3
        assert result["cells_written"] == 6

    async def test_insert_rows(self, module: ExcelModule, wb_path: Path) -> None:
        result = await module._action_insert_rows(
            {"path": str(wb_path), "sheet": "Sheet1", "row": 2, "count": 1}
        )
        assert result["inserted"] is True
        assert result["count"] == 1

    async def test_delete_rows(self, module: ExcelModule, wb_path: Path) -> None:
        result = await module._action_delete_rows(
            {"path": str(wb_path), "sheet": "Sheet1", "row": 3, "count": 1}
        )
        assert result["deleted"] is True

    async def test_insert_columns(self, module: ExcelModule, wb_path: Path) -> None:
        result = await module._action_insert_columns(
            {"path": str(wb_path), "sheet": "Sheet1", "column": 1, "count": 1}
        )
        assert result["inserted"] is True

    async def test_delete_columns(self, module: ExcelModule, wb_path: Path) -> None:
        result = await module._action_delete_columns(
            {"path": str(wb_path), "sheet": "Sheet1", "column": 2, "count": 1}
        )
        assert result["deleted"] is True

    async def test_merge_then_unmerge_cells(self, module: ExcelModule, wb_path: Path) -> None:
        await module._action_merge_cells(
            {"path": str(wb_path), "sheet": "Sheet1", "range": "D5:E6"}
        )
        result = await module._action_unmerge_cells(
            {"path": str(wb_path), "sheet": "Sheet1", "range": "D5:E6"}
        )
        assert result["unmerged"] is True

    async def test_freeze_panes(self, module: ExcelModule, wb_path: Path) -> None:
        result = await module._action_freeze_panes(
            {"path": str(wb_path), "sheet": "Sheet1", "cell": "B2"}
        )
        assert result["freeze_panes"] == "B2"

    async def test_set_column_width(self, module: ExcelModule, wb_path: Path) -> None:
        result = await module._action_set_column_width(
            {"path": str(wb_path), "sheet": "Sheet1", "column": "A", "width": 20}
        )
        assert "A" in result["updated_columns"]

    async def test_set_row_height(self, module: ExcelModule, wb_path: Path) -> None:
        result = await module._action_set_row_height(
            {"path": str(wb_path), "sheet": "Sheet1", "row": 1, "height": 25}
        )
        assert result["row"] == 1
        assert result["height"] == 25


@pytest.mark.integration
class TestSearchAndData:
    async def test_find_replace(self, module: ExcelModule, wb_path: Path) -> None:
        result = await module._action_find_replace(
            {"path": str(wb_path), "sheet": "Sheet1", "find": "Alice", "replace": "Charlie"}
        )
        assert result["replacements"] >= 1

    async def test_remove_duplicates(self, module: ExcelModule, wb_path: Path) -> None:
        # Add a duplicate row
        await module._action_write_range(
            {"path": str(wb_path), "sheet": "Sheet1", "start_cell": "A4",
             "data": [["Alice", 42]]}
        )
        module2 = ExcelModule()
        result = await module2._action_remove_duplicates(
            {"path": str(wb_path), "sheet": "Sheet1", "range": "A1:B4"}
        )
        assert "duplicates_removed" in result

    async def test_apply_formula(self, module: ExcelModule, wb_path: Path) -> None:
        result = await module._action_apply_formula(
            {"path": str(wb_path), "sheet": "Sheet1", "cell": "C1", "formula": "=SUM(B2:B3)"}
        )
        assert result["applied"] is True

    async def test_add_named_range(self, module: ExcelModule, wb_path: Path) -> None:
        result = await module._action_add_named_range(
            {"path": str(wb_path), "sheet": "Sheet1", "name": "DataRange", "range": "A1:B3"}
        )
        assert result["added"] is True

    async def test_add_autofilter(self, module: ExcelModule, wb_path: Path) -> None:
        result = await module._action_add_autofilter(
            {"path": str(wb_path), "sheet": "Sheet1", "range": "A1:B3"}
        )
        assert result["added"] is True


@pytest.mark.integration
class TestFormatting:
    async def test_format_range_fill(self, module: ExcelModule, wb_path: Path) -> None:
        result = await module._action_format_range(
            {
                "path": str(wb_path),
                "sheet": "Sheet1",
                "range": "A1:B1",
                "fill_color": "FFFF00",
                "bold": True,
            }
        )
        assert result["formatted"] is True

    async def test_protect_unprotect_sheet(self, module: ExcelModule, wb_path: Path) -> None:
        protect_result = await module._action_protect_sheet(
            {"path": str(wb_path), "sheet": "Sheet1", "password": "secret"}
        )
        assert protect_result["protected"] is True
        module2 = ExcelModule()
        result = await module2._action_unprotect_sheet(
            {"path": str(wb_path), "sheet": "Sheet1", "password": "secret"}
        )
        assert result["unprotected"] is True

    async def test_set_page_setup(self, module: ExcelModule, wb_path: Path) -> None:
        result = await module._action_set_page_setup(
            {"path": str(wb_path), "sheet": "Sheet1", "orientation": "landscape"}
        )
        assert result["page_setup_applied"] is True


@pytest.mark.integration
class TestExport:
    async def test_export_to_csv(self, module: ExcelModule, wb_path: Path, tmp_path: Path) -> None:
        csv_path = tmp_path / "export.csv"
        result = await module._action_export_to_csv(
            {"path": str(wb_path), "sheet": "Sheet1", "output_path": str(csv_path)}
        )
        assert result["exported"] is True
        assert csv_path.exists()
        content = csv_path.read_text()
        assert "Name" in content
        assert "Alice" in content


@pytest.mark.integration
class TestComments:
    async def test_add_then_delete_comment(self, module: ExcelModule, wb_path: Path) -> None:
        await module._action_add_comment(
            {"path": str(wb_path), "sheet": "Sheet1", "cell": "A1",
             "text": "Test comment", "author": "Tester"}
        )
        module2 = ExcelModule()
        result = await module2._action_delete_comment(
            {"path": str(wb_path), "sheet": "Sheet1", "cell": "A1"}
        )
        assert result["deleted"] is True


# ---------------------------------------------------------------------------
# Extended — copy_range, conditional format, data validation, chart
# ---------------------------------------------------------------------------


@pytest.mark.integration
class TestCopyRange:
    async def test_copy_range_same_sheet(self, module: ExcelModule, wb_path: Path) -> None:
        result = await module._action_copy_range(
            {
                "path": str(wb_path),
                "source_sheet": "Sheet1",
                "source_range": "A1:B2",
                "dest_sheet": "Sheet1",
                "dest_cell": "D1",
            }
        )
        assert result["copied"] is True
        assert result["cells_copied"] == 4

    async def test_copy_range_values_only(self, module: ExcelModule, wb_path: Path) -> None:
        result = await module._action_copy_range(
            {
                "path": str(wb_path),
                "source_sheet": "Sheet1",
                "source_range": "A1:A3",
                "dest_sheet": "Sheet1",
                "dest_cell": "E1",
                "copy_values_only": True,
            }
        )
        assert result["cells_copied"] == 3


@pytest.mark.integration
class TestConditionalFormat:
    async def test_apply_color_scale(self, module: ExcelModule, wb_path: Path) -> None:
        result = await module._action_apply_conditional_format(
            {
                "path": str(wb_path),
                "sheet": "Sheet1",
                "range": "B2:B3",
                "format_type": "color_scale",
                "min_color": "FFFFFF",
                "max_color": "FF0000",
            }
        )
        assert result["applied"] is True
        assert result["format_type"] == "color_scale"

    async def test_apply_data_bar(self, module: ExcelModule, wb_path: Path) -> None:
        result = await module._action_apply_conditional_format(
            {
                "path": str(wb_path),
                "sheet": "Sheet1",
                "range": "B2:B3",
                "format_type": "data_bar",
                "max_color": "638EC6",
            }
        )
        assert result["applied"] is True

    async def test_apply_cell_is(self, module: ExcelModule, wb_path: Path) -> None:
        result = await module._action_apply_conditional_format(
            {
                "path": str(wb_path),
                "sheet": "Sheet1",
                "range": "B2:B3",
                "format_type": "cell_is",
                "operator": "greaterThan",
                "value": 50,
                "fill_color": "FFFF00",
            }
        )
        assert result["applied"] is True

    async def test_apply_formula_format(self, module: ExcelModule, wb_path: Path) -> None:
        result = await module._action_apply_conditional_format(
            {
                "path": str(wb_path),
                "sheet": "Sheet1",
                "range": "A1:B3",
                "format_type": "formula",
                "formula": "=$B1>50",
                "fill_color": "90EE90",
            }
        )
        assert result["applied"] is True

    async def test_apply_icon_set(self, module: ExcelModule, wb_path: Path) -> None:
        result = await module._action_apply_conditional_format(
            {
                "path": str(wb_path),
                "sheet": "Sheet1",
                "range": "B2:B3",
                "format_type": "icon_set",
            }
        )
        assert result["applied"] is True


@pytest.mark.integration
class TestDataValidation:
    async def test_add_list_validation(self, module: ExcelModule, wb_path: Path) -> None:
        result = await module._action_add_data_validation(
            {
                "path": str(wb_path),
                "sheet": "Sheet1",
                "range": "A2:A10",
                "validation_type": "list",
                "formula1": '"Apple,Banana,Cherry"',
            }
        )
        assert result["added"] is True
        assert result["validation_type"] == "list"

    async def test_add_whole_number_validation(self, module: ExcelModule, wb_path: Path) -> None:
        result = await module._action_add_data_validation(
            {
                "path": str(wb_path),
                "sheet": "Sheet1",
                "range": "B2:B5",
                "validation_type": "whole",
                "operator": "between",
                "formula1": "1",
                "formula2": "100",
                "show_error_alert": True,
                "error_message": "Enter a number between 1 and 100",
            }
        )
        assert result["added"] is True


@pytest.mark.integration
class TestCreateChart:
    async def test_create_bar_chart(self, module: ExcelModule, wb_path: Path) -> None:
        result = await module._action_create_chart(
            {
                "path": str(wb_path),
                "sheet": "Sheet1",
                "chart_type": "bar",
                "data_range": "A1:B3",
                "title": "Test Bar Chart",
                "position": "D5",
            }
        )
        assert result["created"] is True
        assert result["chart_type"] in ("bar", "col")

    async def test_create_line_chart(self, module: ExcelModule, wb_path: Path) -> None:
        result = await module._action_create_chart(
            {
                "path": str(wb_path),
                "sheet": "Sheet1",
                "chart_type": "line",
                "data_range": "A1:B3",
                "position": "D15",
            }
        )
        assert result["created"] is True

    async def test_create_pie_chart(self, module: ExcelModule, wb_path: Path) -> None:
        result = await module._action_create_chart(
            {
                "path": str(wb_path),
                "sheet": "Sheet1",
                "chart_type": "pie",
                "data_range": "A1:B3",
                "title": "Pie Chart",
                "position": "H5",
            }
        )
        assert result["created"] is True


@pytest.mark.integration
class TestAutoFilter:
    async def test_add_autofilter(self, module: ExcelModule, wb_path: Path) -> None:
        result = await module._action_add_autofilter(
            {"path": str(wb_path), "sheet": "Sheet1", "range": "A1:B3"}
        )
        assert result["added"] is True
        assert result["sheet"] == "Sheet1"


# ---------------------------------------------------------------------------
# Extended tests for uncovered branches
# ---------------------------------------------------------------------------


@pytest.mark.integration
class TestFormatRangeExtended:
    async def test_format_with_all_font_options(
        self, module: ExcelModule, wb_path: Path
    ) -> None:
        """Covers bold, italic, underline, strikethrough, font_name, font_size, font_color."""
        result = await module._action_format_range(
            {
                "path": str(wb_path),
                "sheet": "Sheet1",
                "range": "A1:B2",
                "bold": True,
                "italic": True,
                "underline": True,
                "strikethrough": True,
                "font_name": "Arial",
                "font_size": 14,
                "font_color": "#FF0000",
            }
        )
        assert result["formatted"] is True
        assert result["cells_formatted"] == 4

    async def test_format_with_fill_none(
        self, module: ExcelModule, wb_path: Path
    ) -> None:
        """Covers fill_type='none' branch."""
        result = await module._action_format_range(
            {
                "path": str(wb_path),
                "sheet": "Sheet1",
                "range": "A1",
                "fill_color": "#FFFFFF",
                "fill_type": "none",
            }
        )
        assert result["formatted"] is True

    async def test_format_with_border_specific_sides(
        self, module: ExcelModule, wb_path: Path
    ) -> None:
        """Covers border with left/right/top/bottom sides (not 'all')."""
        result = await module._action_format_range(
            {
                "path": str(wb_path),
                "sheet": "Sheet1",
                "range": "A1:B1",
                "border_style": "thin",
                "border_color": "#000000",
                "border_sides": ["left", "right"],
            }
        )
        assert result["formatted"] is True

    async def test_format_with_alignment(
        self, module: ExcelModule, wb_path: Path
    ) -> None:
        """Covers alignment branch."""
        result = await module._action_format_range(
            {
                "path": str(wb_path),
                "sheet": "Sheet1",
                "range": "A1",
                "alignment_horizontal": "center",
                "alignment_vertical": "center",
                "wrap_text": True,
            }
        )
        assert result["formatted"] is True

    async def test_format_missing_sheet_raises(
        self, module: ExcelModule, wb_path: Path
    ) -> None:
        with pytest.raises(KeyError):
            await module._action_format_range(
                {
                    "path": str(wb_path),
                    "sheet": "NoSuchSheet",
                    "range": "A1",
                }
            )

    async def test_format_with_number_format(
        self, module: ExcelModule, wb_path: Path
    ) -> None:
        result = await module._action_format_range(
            {
                "path": str(wb_path),
                "sheet": "Sheet1",
                "range": "B2:B3",
                "number_format": "0.00",
            }
        )
        assert result["formatted"] is True


@pytest.mark.integration
class TestSetColumnWidthExtended:
    async def test_set_column_range_width(
        self, module: ExcelModule, wb_path: Path
    ) -> None:
        """Covers column range parsing (A:C)."""
        result = await module._action_set_column_width(
            {"path": str(wb_path), "sheet": "Sheet1", "column": "A:B", "width": 20.0}
        )
        assert "A" in result["updated_columns"]
        assert "B" in result["updated_columns"]

    async def test_set_column_auto_fit(
        self, module: ExcelModule, wb_path: Path
    ) -> None:
        """Covers auto_fit branch."""
        result = await module._action_set_column_width(
            {
                "path": str(wb_path),
                "sheet": "Sheet1",
                "column": "A",
                "auto_fit": True,
                "width": None,
            }
        )
        assert "A" in result["updated_columns"]

    async def test_set_column_width_missing_sheet(
        self, module: ExcelModule, wb_path: Path
    ) -> None:
        with pytest.raises(KeyError):
            await module._action_set_column_width(
                {"path": str(wb_path), "sheet": "Missing", "column": "A", "width": 10}
            )


@pytest.mark.integration
class TestSetRowHeightExtended:
    async def test_set_row_height_missing_sheet(
        self, module: ExcelModule, wb_path: Path
    ) -> None:
        with pytest.raises(KeyError):
            await module._action_set_row_height(
                {"path": str(wb_path), "sheet": "Missing", "row": 1, "height": 20}
            )


@pytest.mark.integration
class TestFindReplaceExtended:
    async def test_find_replace_whole_cell(
        self, module: ExcelModule, wb_path: Path
    ) -> None:
        """Covers whole_cell=True path."""
        result = await module._action_find_replace(
            {
                "path": str(wb_path),
                "sheet": "Sheet1",
                "find": "Alice",
                "replace": "Alicia",
                "whole_cell": True,
                "case_sensitive": True,
            }
        )
        assert result["replacements"] == 1

    async def test_find_replace_case_sensitive_partial(
        self, module: ExcelModule, wb_path: Path
    ) -> None:
        """Covers case_sensitive=True with partial match."""
        result = await module._action_find_replace(
            {
                "path": str(wb_path),
                "sheet": "Sheet1",
                "find": "ob",
                "replace": "OB",
                "case_sensitive": True,
            }
        )
        assert result["replacements"] >= 1


@pytest.mark.integration
class TestRemoveDuplicatesExtended:
    async def test_remove_duplicates_keep_last(
        self, module: ExcelModule, tmp_path: Path
    ) -> None:
        """Covers keep='last' branch."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws["A1"] = "Name"
        ws["A2"] = "Alice"
        ws["A3"] = "Alice"  # duplicate
        ws["A4"] = "Bob"
        path = tmp_path / "dups.xlsx"
        wb.save(str(path))

        result = await module._action_remove_duplicates(
            {"path": str(path), "sheet": "Sheet1", "range": "A1:A4", "keep": "last"}
        )
        assert result["duplicates_removed"] == 1

    async def test_remove_duplicates_missing_sheet(
        self, module: ExcelModule, wb_path: Path
    ) -> None:
        with pytest.raises(KeyError):
            await module._action_remove_duplicates(
                {"path": str(wb_path), "sheet": "Missing", "range": "A1:A3"}
            )


@pytest.mark.integration
class TestFreezePanesExtended:
    async def test_freeze_panes_missing_sheet(
        self, module: ExcelModule, wb_path: Path
    ) -> None:
        with pytest.raises(KeyError):
            await module._action_freeze_panes(
                {"path": str(wb_path), "sheet": "Missing", "cell": "B2"}
            )


@pytest.mark.integration
class TestPageSetupExtended:
    async def test_page_setup_with_scale_and_print_area(
        self, module: ExcelModule, wb_path: Path
    ) -> None:
        """Covers scale, print_area, print_title_rows, print_title_cols branches."""
        result = await module._action_set_page_setup(
            {
                "path": str(wb_path),
                "sheet": "Sheet1",
                "scale": 90,
                "print_area": "A1:B10",
                "print_title_rows": "1:1",
                "print_title_cols": "A:A",
            }
        )
        assert result["page_setup_applied"] is True


@pytest.mark.integration
class TestGetWorkbookInfoExtended:
    async def test_include_named_ranges(
        self, module: ExcelModule, tmp_path: Path
    ) -> None:
        """Covers include_named_ranges branch in get_workbook_info."""
        from openpyxl.workbook.defined_name import DefinedName

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws["A1"] = 1
        dn = DefinedName(name="MyRange", attr_text="'Sheet1'!$A$1")
        wb.defined_names.add(dn)
        path = tmp_path / "named.xlsx"
        wb.save(str(path))

        result = await module._action_get_workbook_info(
            {"path": str(path), "include_named_ranges": True}
        )
        assert "defined_names" in result
        assert any(n["name"] == "MyRange" for n in result["defined_names"])


@pytest.mark.integration
class TestGetSheetInfoExtended:
    async def test_include_dimensions_and_merged(
        self, module: ExcelModule, tmp_path: Path
    ) -> None:
        """Covers include_dimensions and include_merged_cells branches."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws.merge_cells("A1:B2")
        path = tmp_path / "merged.xlsx"
        wb.save(str(path))

        result = await module._action_get_sheet_info(
            {
                "path": str(path),
                "sheet": "Sheet1",
                "include_dimensions": True,
                "include_merged_cells": True,
            }
        )
        assert "dimensions" in result
        assert "merged_cells" in result

    async def test_get_sheet_info_missing_sheet(
        self, module: ExcelModule, wb_path: Path
    ) -> None:
        with pytest.raises(KeyError):
            await module._action_get_sheet_info(
                {"path": str(wb_path), "sheet": "NoSuchSheet"}
            )


@pytest.mark.integration
class TestCopySheetExtended:
    async def test_copy_sheet_with_position(
        self, module: ExcelModule, wb_path: Path
    ) -> None:
        """Covers position parameter in copy_sheet."""
        result = await module._action_copy_sheet(
            {
                "path": str(wb_path),
                "source_sheet": "Sheet1",
                "new_name": "Sheet1_copy",
                "position": 0,
            }
        )
        assert result["copied"] is True

    async def test_copy_sheet_missing_source(
        self, module: ExcelModule, wb_path: Path
    ) -> None:
        with pytest.raises(KeyError):
            await module._action_copy_sheet(
                {
                    "path": str(wb_path),
                    "source_sheet": "NoSuch",
                    "new_name": "Copy",
                }
            )


@pytest.mark.integration
class TestNamedRangeExtended:
    async def test_add_named_range_sheet_scope(
        self, module: ExcelModule, wb_path: Path
    ) -> None:
        """Covers scope='sheet' branch."""
        result = await module._action_add_named_range(
            {
                "path": str(wb_path),
                "name": "LocalRange",
                "sheet": "Sheet1",
                "range": "A1:B2",
                "scope": "sheet",
            }
        )
        assert result["added"] is True
        assert result["scope"] == "sheet"

    async def test_add_named_range_missing_sheet(
        self, module: ExcelModule, wb_path: Path
    ) -> None:
        with pytest.raises(KeyError):
            await module._action_add_named_range(
                {
                    "path": str(wb_path),
                    "name": "Bad",
                    "sheet": "NoSuch",
                    "range": "A1",
                    "scope": "workbook",
                }
            )


@pytest.mark.integration
class TestInsertImageExcel:
    async def test_insert_image(
        self, module: ExcelModule, wb_path: Path, tmp_path: Path
    ) -> None:
        """Covers _action_insert_image (lines 1195-1220)."""
        import struct
        import zlib

        def _make_tiny_png() -> bytes:
            sig = b"\x89PNG\r\n\x1a\n"
            ihdr_data = struct.pack(">IIBBBBB", 1, 1, 8, 2, 0, 0, 0)
            ihdr_crc = struct.pack(">I", zlib.crc32(b"IHDR" + ihdr_data) & 0xFFFFFFFF)
            ihdr = struct.pack(">I", 13) + b"IHDR" + ihdr_data + ihdr_crc
            raw = b"\x00\xff\xff\xff"
            compressed = zlib.compress(raw)
            idat_crc = struct.pack(">I", zlib.crc32(b"IDAT" + compressed) & 0xFFFFFFFF)
            idat = struct.pack(">I", len(compressed)) + b"IDAT" + compressed + idat_crc
            iend_crc = struct.pack(">I", zlib.crc32(b"IEND") & 0xFFFFFFFF)
            iend = struct.pack(">I", 0) + b"IEND" + iend_crc
            return sig + ihdr + idat + iend

        img_path = tmp_path / "tiny.png"
        img_path.write_bytes(_make_tiny_png())

        result = await module._action_insert_image(
            {
                "path": str(wb_path),
                "sheet": "Sheet1",
                "image_path": str(img_path),
                "cell": "D1",
                "width": 50,
                "height": 50,
            }
        )
        assert result["inserted"] is True

    async def test_insert_image_missing_sheet(
        self, module: ExcelModule, wb_path: Path, tmp_path: Path
    ) -> None:
        img_path = tmp_path / "fake.png"
        img_path.write_bytes(b"PNG")
        with pytest.raises((KeyError, Exception)):
            await module._action_insert_image(
                {
                    "path": str(wb_path),
                    "sheet": "NoSuch",
                    "image_path": str(img_path),
                    "cell": "A1",
                }
            )


@pytest.mark.integration
class TestExportToPdfExtended:
    async def test_export_to_pdf_no_libreoffice_raises(
        self, module: ExcelModule, wb_path: Path, tmp_path: Path
    ) -> None:
        """use_libreoffice=False raises RuntimeError."""
        with pytest.raises(RuntimeError, match="LibreOffice"):
            await module._action_export_to_pdf(
                {
                    "path": str(wb_path),
                    "output_path": str(tmp_path / "out.pdf"),
                    "use_libreoffice": False,
                }
            )

    async def test_export_to_pdf_with_libreoffice_mocked(
        self, module: ExcelModule, wb_path: Path, tmp_path: Path
    ) -> None:
        """Covers the LibreOffice subprocess path."""
        import subprocess
        from unittest.mock import MagicMock, patch

        out_pdf = tmp_path / "out.pdf"
        # LibreOffice generates a .pdf with same stem in outdir
        expected_pdf = tmp_path / (wb_path.stem + ".pdf")

        def _mock_run(cmd, **kwargs):
            # Create the PDF that LibreOffice would generate
            expected_pdf.write_bytes(b"fake pdf content")
            mock_result = MagicMock()
            mock_result.returncode = 0
            mock_result.stderr = ""
            return mock_result

        with patch("subprocess.run", side_effect=_mock_run):
            result = await module._action_export_to_pdf(
                {
                    "path": str(wb_path),
                    "output_path": str(out_pdf),
                    "use_libreoffice": True,
                }
            )
        assert result["exported"] is True

    async def test_export_to_pdf_libreoffice_fails(
        self, module: ExcelModule, wb_path: Path, tmp_path: Path
    ) -> None:
        """Covers LibreOffice returncode != 0 → raises RuntimeError."""
        from unittest.mock import MagicMock, patch

        mock_result = MagicMock()
        mock_result.returncode = 1
        mock_result.stderr = "conversion failed"

        with patch("subprocess.run", return_value=mock_result):
            with pytest.raises(RuntimeError, match="LibreOffice"):
                await module._action_export_to_pdf(
                    {
                        "path": str(wb_path),
                        "output_path": str(tmp_path / "out.pdf"),
                        "use_libreoffice": True,
                    }
                )


@pytest.mark.integration
class TestScatterChart:
    async def test_create_scatter_chart(
        self, module: ExcelModule, tmp_path: Path
    ) -> None:
        """Covers scatter chart series building (lines 1154-1162)."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        for i in range(1, 6):
            ws.cell(row=i, column=1, value=i)
            ws.cell(row=i, column=2, value=i * 2)
        path = tmp_path / "scatter.xlsx"
        wb.save(str(path))

        result = await module._action_create_chart(
            {
                "path": str(path),
                "sheet": "Sheet1",
                "chart_type": "scatter",
                "data_range": "A1:B5",
                "position": "D1",
            }
        )
        assert result["created"] is True
        assert result["chart_type"] == "scatter"

    async def test_create_chart_with_axis_titles_and_no_legend(
        self, module: ExcelModule, tmp_path: Path
    ) -> None:
        """Covers x_axis_title, y_axis_title, has_legend=False branches."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        for i in range(1, 6):
            ws.cell(row=i, column=1, value=i)
            ws.cell(row=i, column=2, value=i * 2)
        path = tmp_path / "line_chart.xlsx"
        wb.save(str(path))

        result = await module._action_create_chart(
            {
                "path": str(path),
                "sheet": "Sheet1",
                "chart_type": "line",
                "data_range": "A1:B5",
                "position": "D1",
                "x_axis_title": "X Axis",
                "y_axis_title": "Y Axis",
                "has_legend": False,
                "title": "My Line Chart",
            }
        )
        assert result["created"] is True

    async def test_create_chart_col_type(
        self, module: ExcelModule, tmp_path: Path
    ) -> None:
        """Covers 'col' chart type."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        for i in range(1, 4):
            ws.cell(row=i, column=1, value=i)
        path = tmp_path / "col_chart.xlsx"
        wb.save(str(path))

        result = await module._action_create_chart(
            {
                "path": str(path),
                "sheet": "Sheet1",
                "chart_type": "col",
                "data_range": "A1:A3",
                "position": "C1",
            }
        )
        assert result["created"] is True

    async def test_create_chart_with_series_labels(
        self, module: ExcelModule, tmp_path: Path
    ) -> None:
        """Covers series_labels=True with categories (lines 1163-1171)."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws["A1"] = "Category"
        ws["B1"] = "Values"
        for i in range(2, 6):
            ws.cell(row=i, column=1, value=f"Cat{i}")
            ws.cell(row=i, column=2, value=i * 10)
        path = tmp_path / "labels_chart.xlsx"
        wb.save(str(path))

        result = await module._action_create_chart(
            {
                "path": str(path),
                "sheet": "Sheet1",
                "chart_type": "bar",
                "data_range": "A1:B5",
                "position": "D1",
                "series_labels": True,
            }
        )
        assert result["created"] is True

    async def test_create_chart_missing_sheet(
        self, module: ExcelModule, wb_path: Path
    ) -> None:
        with pytest.raises(KeyError):
            await module._action_create_chart(
                {
                    "path": str(wb_path),
                    "sheet": "NoSuch",
                    "chart_type": "bar",
                    "data_range": "A1:B3",
                    "position": "D1",
                }
            )


@pytest.mark.integration
class TestCellValueHelpers:
    async def test_cell_value_datetime(
        self, module: ExcelModule, tmp_path: Path
    ) -> None:
        """Covers datetime branch in _cell_to_value."""
        import datetime

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws["A1"] = datetime.datetime(2024, 1, 15, 10, 30, 0)
        ws["A2"] = datetime.date(2024, 1, 15)
        path = tmp_path / "dates.xlsx"
        wb.save(str(path))

        result = await module._action_read_range(
            {"path": str(path), "sheet": "Sheet1", "range": "A1:A2"}
        )
        # Should return ISO format strings
        assert "2024-01-15" in str(result["data"][0][0])

    async def test_parse_range_single_cell(
        self, module: ExcelModule, wb_path: Path
    ) -> None:
        """Covers single-cell range parsing (no ':' in range)."""
        result = await module._action_read_range(
            {"path": str(wb_path), "sheet": "Sheet1", "range": "A1"}
        )
        assert result["row_count"] == 1

    async def test_hex_to_rgb_invalid_raises(self, module: ExcelModule) -> None:
        """Covers ValueError in _hex_to_rgb."""
        with pytest.raises(ValueError, match="Invalid hex color"):
            module._hex_to_rgb("XYZ")


@pytest.mark.integration
class TestExportCsvExtended:
    async def test_export_csv_no_header(
        self, module: ExcelModule, wb_path: Path, tmp_path: Path
    ) -> None:
        """Covers include_header=False branch (line 1280)."""
        out_csv = tmp_path / "out.csv"
        result = await module._action_export_to_csv(
            {
                "path": str(wb_path),
                "sheet": "Sheet1",
                "output_path": str(out_csv),
                "include_header": False,
            }
        )
        assert result["exported"] is True
        # Should have 2 data rows (A2, A3) skipping header
        assert result["rows_written"] == 2

    async def test_export_csv_missing_sheet(
        self, module: ExcelModule, wb_path: Path, tmp_path: Path
    ) -> None:
        with pytest.raises(KeyError):
            await module._action_export_to_csv(
                {
                    "path": str(wb_path),
                    "sheet": "NoSuch",
                    "output_path": str(tmp_path / "out.csv"),
                }
            )


@pytest.mark.integration
class TestSaveWithOutputPath:
    async def test_save_workbook_to_different_path(
        self, module: ExcelModule, wb_path: Path, tmp_path: Path
    ) -> None:
        """Covers the src_key != target branch in _save_wb (line 136)."""
        # Load workbook into cache
        await module._action_open_workbook({"path": str(wb_path)})
        out_path = tmp_path / "copy.xlsx"
        result = await module._action_save_workbook(
            {"path": str(wb_path), "output_path": str(out_path)}
        )
        assert result["saved"] is True
        assert out_path.exists()


@pytest.mark.integration
class TestMiscKeyErrors:
    async def test_protect_sheet_missing_sheet(
        self, module: ExcelModule, wb_path: Path
    ) -> None:
        with pytest.raises(KeyError):
            await module._action_protect_sheet(
                {"path": str(wb_path), "sheet": "Missing"}
            )

    async def test_unprotect_sheet_missing_sheet(
        self, module: ExcelModule, wb_path: Path
    ) -> None:
        with pytest.raises(KeyError):
            await module._action_unprotect_sheet(
                {"path": str(wb_path), "sheet": "Missing"}
            )

    async def test_page_setup_missing_sheet(
        self, module: ExcelModule, wb_path: Path
    ) -> None:
        with pytest.raises(KeyError):
            await module._action_set_page_setup(
                {"path": str(wb_path), "sheet": "Missing"}
            )

    async def test_add_comment_missing_sheet(
        self, module: ExcelModule, wb_path: Path
    ) -> None:
        with pytest.raises(KeyError):
            await module._action_add_comment(
                {
                    "path": str(wb_path),
                    "sheet": "Missing",
                    "cell": "A1",
                    "text": "hello",
                    "author": "me",
                }
            )

    async def test_delete_comment_missing_sheet(
        self, module: ExcelModule, wb_path: Path
    ) -> None:
        with pytest.raises(KeyError):
            await module._action_delete_comment(
                {"path": str(wb_path), "sheet": "Missing", "cell": "A1"}
            )

    async def test_autofilter_missing_sheet(
        self, module: ExcelModule, wb_path: Path
    ) -> None:
        with pytest.raises(KeyError):
            await module._action_add_autofilter(
                {"path": str(wb_path), "sheet": "Missing", "range": "A1:B3"}
            )
