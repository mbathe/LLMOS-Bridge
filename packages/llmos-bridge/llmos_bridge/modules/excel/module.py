"""Excel module — full openpyxl-powered spreadsheet automation.

All blocking I/O is offloaded to a thread via ``asyncio.to_thread``.
Workbooks are cached in ``_wb_cache`` (keyed by resolved path string)
to avoid redundant disk reads within a session.

Dependencies (lazy-loaded):
    openpyxl >= 3.1   — install with ``pip install openpyxl``
"""

from __future__ import annotations

import asyncio
import csv
import subprocess
import threading
from pathlib import Path
from typing import Any

try:
    from llmos_bridge.exceptions import ModuleLoadError
except ImportError:
    ModuleLoadError = ImportError  # type: ignore[misc,assignment]

from llmos_bridge.modules.base import BaseModule, Platform
from llmos_bridge.modules.manifest import ActionSpec, ModuleManifest, ParamSpec
from llmos_bridge.protocol.params.excel import (
    AddAutoFilterParams,
    AddCommentParams,
    AddDataValidationParams,
    AddNamedRangeParams,
    ApplyConditionalFormatParams,
    ApplyFormulaParams,
    CloseWorkbookParams,
    CopyRangeParams,
    CopySheetParams,
    CreateChartParams,
    CreateSheetParams,
    CreateWorkbookParams,
    DeleteColumnsParams,
    DeleteCommentParams,
    DeleteRowsParams,
    DeleteSheetParams,
    ExportToCsvParams,
    ExportToPdfParams,
    FindReplaceParams,
    FormatRangeParams,
    FreezePanesParams,
    GetSheetInfoParams,
    GetWorkbookInfoParams,
    InsertColumnsParams,
    InsertImageParams,
    InsertRowsParams,
    ListSheetsParams,
    MergeCellsParams,
    OpenWorkbookParams,
    ProtectSheetParams,
    ReadCellParams,
    ReadRangeParams,
    RemoveDuplicatesParams,
    RenameSheetParams,
    SaveWorkbookParams,
    SetColumnWidthParams,
    SetPageSetupParams,
    SetRowHeightParams,
    UnmergeCellsParams,
    UnprotectSheetParams,
    WriteCellParams,
    WriteRangeParams,
)

MODULE_ID = "excel"
VERSION = "1.0.0"
SUPPORTED_PLATFORMS = [Platform.ALL]


class ExcelModule(BaseModule):
    """Full-featured Excel spreadsheet automation using openpyxl."""

    MODULE_ID = MODULE_ID
    VERSION = VERSION
    SUPPORTED_PLATFORMS = SUPPORTED_PLATFORMS

    def __init__(self) -> None:
        self._wb_cache: dict[str, Any] = {}
        self._path_locks: dict[str, threading.Lock] = {}
        self._meta_lock = threading.Lock()
        super().__init__()

    def _get_path_lock(self, path: str) -> threading.Lock:
        """Return (or create) a per-file threading.Lock for concurrent access control."""
        resolved = str(Path(path).resolve())
        with self._meta_lock:
            if resolved not in self._path_locks:
                self._path_locks[resolved] = threading.Lock()
            return self._path_locks[resolved]

    # ------------------------------------------------------------------
    # Dependency check
    # ------------------------------------------------------------------

    def _check_dependencies(self) -> None:
        try:
            import openpyxl  # noqa: F401
        except ImportError as exc:
            raise ModuleLoadError(
                module_id=MODULE_ID,
                reason="openpyxl is required but not installed. Run: pip install openpyxl",
            ) from exc

    # ------------------------------------------------------------------
    # Internal helpers
    # ------------------------------------------------------------------

    def _get_wb(
        self,
        path: str,
        read_only: bool = False,
        data_only: bool = True,
        keep_vba: bool = False,
    ) -> Any:
        """Return cached workbook or load from disk."""
        import openpyxl

        resolved = str(Path(path).resolve())
        if resolved in self._wb_cache:
            return self._wb_cache[resolved]
        wb = openpyxl.load_workbook(
            resolved,
            read_only=read_only,
            data_only=data_only,
            keep_vba=keep_vba,
        )
        if not read_only:
            self._wb_cache[resolved] = wb
        return wb

    def _save_wb(self, path: str, wb: Any, output_path: str | None = None) -> str:
        """Save workbook to *output_path* (or *path* if None). Returns saved path."""
        target = str(Path(output_path).resolve()) if output_path else str(Path(path).resolve())
        Path(target).parent.mkdir(parents=True, exist_ok=True)
        wb.save(target)
        # Update cache: if saving to a new location, add that key too.
        src_key = str(Path(path).resolve())
        self._wb_cache[target] = wb
        if src_key != target and src_key in self._wb_cache:
            # Keep original cache entry up-to-date (same object).
            self._wb_cache[src_key] = wb
        return target

    def _parse_range(
        self, range_str: str, ws: Any
    ) -> tuple[int, int, int, int]:
        """Parse ``'A1:B10'`` or ``'auto'`` → (min_row, max_row, min_col, max_col)."""
        from openpyxl.utils import column_index_from_string

        if range_str.lower() == "auto":
            return ws.min_row, ws.max_row, ws.min_column, ws.max_column

        if ":" in range_str:
            start_cell, end_cell = range_str.split(":", 1)
        else:
            # Single cell: return a 1×1 range.
            start_cell = end_cell = range_str

        def _split_cell(cell: str) -> tuple[int, int]:
            # Split "AB12" → ("AB", "12")
            col_str = "".join(c for c in cell if c.isalpha())
            row_str = "".join(c for c in cell if c.isdigit())
            return int(row_str), column_index_from_string(col_str)

        min_row, min_col = _split_cell(start_cell.strip())
        max_row, max_col = _split_cell(end_cell.strip())
        return min_row, max_row, min_col, max_col

    def _cell_to_value(self, cell: Any) -> Any:
        """Convert a cell value to a JSON-safe Python type."""
        import datetime

        v = cell.value
        if isinstance(v, (datetime.datetime, datetime.date, datetime.time)):
            return v.isoformat()
        if isinstance(v, (int, float, bool, str)) or v is None:
            return v
        return str(v)

    def _hex_to_rgb(self, hex_color: str) -> tuple[int, int, int]:
        """Convert ``'FF0000'`` (or ``'#FF0000'``) to ``(255, 0, 0)``."""
        hex_color = hex_color.lstrip("#")
        if len(hex_color) == 6:
            r = int(hex_color[0:2], 16)
            g = int(hex_color[2:4], 16)
            b = int(hex_color[4:6], 16)
            return r, g, b
        raise ValueError(f"Invalid hex color: {hex_color!r}")

    # ------------------------------------------------------------------
    # Workbook lifecycle
    # ------------------------------------------------------------------

    async def _action_create_workbook(self, params: dict[str, Any]) -> dict[str, Any]:
        p = CreateWorkbookParams.model_validate(params)

        def _create() -> dict[str, Any]:
            with self._get_path_lock(p.path):
                import openpyxl

                wb = openpyxl.Workbook()
                # Rename the default sheet.
                ws = wb.active
                ws.title = p.sheet_name

                out = Path(p.path).resolve()
                out.parent.mkdir(parents=True, exist_ok=True)
                wb.save(str(out))
                self._wb_cache[str(out)] = wb
                return {
                    "path": str(out),
                    "created": True,
                    "sheet_names": wb.sheetnames,
                }

        return await asyncio.to_thread(_create)

    async def _action_open_workbook(self, params: dict[str, Any]) -> dict[str, Any]:
        p = OpenWorkbookParams.model_validate(params)

        def _load() -> dict[str, Any]:
            with self._get_path_lock(p.path):
                wb = self._get_wb(
                    p.path,
                    read_only=p.read_only,
                    data_only=p.data_only,
                    keep_vba=p.keep_vba,
                )
                return {
                    "path": str(Path(p.path).resolve()),
                    "sheet_names": wb.sheetnames,
                    "active_sheet": wb.active.title if wb.active else None,
                    "read_only": p.read_only,
                    "data_only": p.data_only,
                }

        return await asyncio.to_thread(_load)

    async def _action_close_workbook(self, params: dict[str, Any]) -> dict[str, Any]:
        p = CloseWorkbookParams.model_validate(params)

        def _close() -> dict[str, Any]:
            with self._get_path_lock(p.path):
                resolved = str(Path(p.path).resolve())
                wb = self._wb_cache.get(resolved)
                if wb is None:
                    return {"closed": True, "was_cached": False}
                if p.save:
                    self._save_wb(p.path, wb)
                del self._wb_cache[resolved]
                return {"closed": True, "was_cached": True, "saved": p.save}

        return await asyncio.to_thread(_close)

    async def _action_save_workbook(self, params: dict[str, Any]) -> dict[str, Any]:
        p = SaveWorkbookParams.model_validate(params)

        def _save() -> dict[str, Any]:
            with self._get_path_lock(p.path):
                wb = self._get_wb(p.path, data_only=False)
                saved_path = self._save_wb(p.path, wb, p.output_path)
                return {"saved": True, "path": saved_path}

        return await asyncio.to_thread(_save)

    async def _action_get_workbook_info(self, params: dict[str, Any]) -> dict[str, Any]:
        p = GetWorkbookInfoParams.model_validate(params)

        def _info() -> dict[str, Any]:
            with self._get_path_lock(p.path):
                wb = self._get_wb(p.path)
                result: dict[str, Any] = {
                    "path": str(Path(p.path).resolve()),
                    "title": wb.properties.title,
                    "sheet_names": wb.sheetnames,
                    "active_sheet": wb.active.title if wb.active else None,
                }
                if p.include_named_ranges or p.include_defined_names:
                    named: list[dict[str, Any]] = []
                    for dn in wb.defined_names.values():
                        named.append({"name": dn.name, "value": dn.attr_text})
                    result["defined_names"] = named
                return result

        return await asyncio.to_thread(_info)

    # ------------------------------------------------------------------
    # Sheet management
    # ------------------------------------------------------------------

    async def _action_list_sheets(self, params: dict[str, Any]) -> dict[str, Any]:
        p = ListSheetsParams.model_validate(params)

        def _list() -> dict[str, Any]:
            with self._get_path_lock(p.path):
                wb = self._get_wb(p.path)
                return {"sheet_names": wb.sheetnames, "count": len(wb.sheetnames)}

        return await asyncio.to_thread(_list)

    async def _action_create_sheet(self, params: dict[str, Any]) -> dict[str, Any]:
        p = CreateSheetParams.model_validate(params)

        def _create() -> dict[str, Any]:
            with self._get_path_lock(p.path):
                wb = self._get_wb(p.path, data_only=False)
                ws = wb.create_sheet(title=p.name, index=p.position)
                self._save_wb(p.path, wb)
                return {"created": True, "sheet_name": ws.title, "position": wb.sheetnames.index(ws.title)}

        return await asyncio.to_thread(_create)

    async def _action_delete_sheet(self, params: dict[str, Any]) -> dict[str, Any]:
        p = DeleteSheetParams.model_validate(params)

        def _delete() -> dict[str, Any]:
            with self._get_path_lock(p.path):
                wb = self._get_wb(p.path, data_only=False)
                if p.name not in wb.sheetnames:
                    raise KeyError(f"Sheet '{p.name}' not found.")
                del wb[p.name]
                self._save_wb(p.path, wb)
                return {"deleted": True, "sheet_name": p.name}

        return await asyncio.to_thread(_delete)

    async def _action_rename_sheet(self, params: dict[str, Any]) -> dict[str, Any]:
        p = RenameSheetParams.model_validate(params)

        def _rename() -> dict[str, Any]:
            with self._get_path_lock(p.path):
                wb = self._get_wb(p.path, data_only=False)
                if p.old_name not in wb.sheetnames:
                    raise KeyError(f"Sheet '{p.old_name}' not found.")
                wb[p.old_name].title = p.new_name
                self._save_wb(p.path, wb)
                return {"renamed": True, "old_name": p.old_name, "new_name": p.new_name}

        return await asyncio.to_thread(_rename)

    async def _action_copy_sheet(self, params: dict[str, Any]) -> dict[str, Any]:
        p = CopySheetParams.model_validate(params)

        def _copy() -> dict[str, Any]:
            with self._get_path_lock(p.path):
                wb = self._get_wb(p.path, data_only=False)
                if p.source_sheet not in wb.sheetnames:
                    raise KeyError(f"Source sheet '{p.source_sheet}' not found.")
                ws_copy = wb.copy_worksheet(wb[p.source_sheet])
                ws_copy.title = p.new_name
                if p.position is not None:
                    # Move to desired position.
                    wb.move_sheet(ws_copy, offset=p.position - wb.sheetnames.index(ws_copy.title))
                self._save_wb(p.path, wb)
                return {
                    "copied": True,
                    "source_sheet": p.source_sheet,
                    "new_name": ws_copy.title,
                    "position": wb.sheetnames.index(ws_copy.title),
                }

        return await asyncio.to_thread(_copy)

    async def _action_get_sheet_info(self, params: dict[str, Any]) -> dict[str, Any]:
        p = GetSheetInfoParams.model_validate(params)

        def _info() -> dict[str, Any]:
            with self._get_path_lock(p.path):
                wb = self._get_wb(p.path)
                if p.sheet not in wb.sheetnames:
                    raise KeyError(f"Sheet '{p.sheet}' not found.")
                ws = wb[p.sheet]
                result: dict[str, Any] = {
                    "sheet_name": ws.title,
                    "freeze_panes": str(ws.freeze_panes) if ws.freeze_panes else None,
                }
                if p.include_dimensions:
                    result["dimensions"] = {
                        "min_row": ws.min_row,
                        "max_row": ws.max_row,
                        "min_column": ws.min_column,
                        "max_column": ws.max_column,
                        "used_range": ws.dimensions,
                    }
                if p.include_merged_cells:
                    result["merged_cells"] = [str(mc) for mc in ws.merged_cells.ranges]
                return result

        return await asyncio.to_thread(_info)

    async def _action_protect_sheet(self, params: dict[str, Any]) -> dict[str, Any]:
        p = ProtectSheetParams.model_validate(params)

        def _protect() -> dict[str, Any]:
            with self._get_path_lock(p.path):
                wb = self._get_wb(p.path, data_only=False)
                if p.sheet not in wb.sheetnames:
                    raise KeyError(f"Sheet '{p.sheet}' not found.")
                ws = wb[p.sheet]
                ws.protection.sheet = True
                if p.password:
                    ws.protection.set_password(p.password)
                ws.protection.selectLockedCells = not p.allow_select_locked
                ws.protection.selectUnlockedCells = not p.allow_select_unlocked
                ws.protection.formatCells = not p.allow_format_cells
                ws.protection.sort = not p.allow_sort
                ws.protection.autoFilter = not p.allow_auto_filter
                ws.protection.insertRows = not p.allow_insert_rows
                ws.protection.deleteRows = not p.allow_delete_rows
                self._save_wb(p.path, wb)
                return {"protected": True, "sheet": p.sheet}

        return await asyncio.to_thread(_protect)

    async def _action_unprotect_sheet(self, params: dict[str, Any]) -> dict[str, Any]:
        p = UnprotectSheetParams.model_validate(params)

        def _unprotect() -> dict[str, Any]:
            with self._get_path_lock(p.path):
                wb = self._get_wb(p.path, data_only=False)
                if p.sheet not in wb.sheetnames:
                    raise KeyError(f"Sheet '{p.sheet}' not found.")
                ws = wb[p.sheet]
                ws.protection.sheet = False
                ws.protection.set_password("")
                self._save_wb(p.path, wb)
                return {"unprotected": True, "sheet": p.sheet}

        return await asyncio.to_thread(_unprotect)

    async def _action_set_page_setup(self, params: dict[str, Any]) -> dict[str, Any]:
        p = SetPageSetupParams.model_validate(params)

        def _setup() -> dict[str, Any]:
            with self._get_path_lock(p.path):
                wb = self._get_wb(p.path, data_only=False)
                if p.sheet not in wb.sheetnames:
                    raise KeyError(f"Sheet '{p.sheet}' not found.")
                ws = wb[p.sheet]
                ws.page_setup.orientation = p.orientation
                ws.page_setup.paperSize = p.paper_size
                ws.page_setup.fitToPage = p.fit_to_page
                ws.page_setup.fitToWidth = p.fit_to_width
                ws.page_setup.fitToHeight = p.fit_to_height
                if p.scale is not None:
                    ws.page_setup.scale = p.scale
                ws.page_margins.top = p.top_margin
                ws.page_margins.bottom = p.bottom_margin
                ws.page_margins.left = p.left_margin
                ws.page_margins.right = p.right_margin
                ws.page_margins.header = p.header_margin
                ws.page_margins.footer = p.footer_margin
                if p.print_area:
                    ws.print_area = p.print_area
                if p.print_title_rows:
                    ws.print_title_rows = p.print_title_rows
                if p.print_title_cols:
                    ws.print_title_cols = p.print_title_cols
                self._save_wb(p.path, wb)
                return {"page_setup_applied": True, "sheet": p.sheet}

        return await asyncio.to_thread(_setup)

    # ------------------------------------------------------------------
    # Cell & range operations
    # ------------------------------------------------------------------

    async def _action_read_cell(self, params: dict[str, Any]) -> dict[str, Any]:
        p = ReadCellParams.model_validate(params)

        def _read() -> dict[str, Any]:
            with self._get_path_lock(p.path):
                wb = self._get_wb(p.path, data_only=True)
                if p.sheet not in wb.sheetnames:
                    raise KeyError(f"Sheet '{p.sheet}' not found.")
                ws = wb[p.sheet]
                cell = ws[p.cell]
                comment_text = cell.comment.text if cell.comment else None
                return {
                    "cell": p.cell,
                    "value": self._cell_to_value(cell),
                    "data_type": cell.data_type,
                    "formula": cell.value if isinstance(cell.value, str) and cell.value.startswith("=") else None,
                    "comment": comment_text,
                    "number_format": cell.number_format,
                }

        return await asyncio.to_thread(_read)

    async def _action_write_cell(self, params: dict[str, Any]) -> dict[str, Any]:
        p = WriteCellParams.model_validate(params)

        def _write() -> dict[str, Any]:
            with self._get_path_lock(p.path):
                wb = self._get_wb(p.path, data_only=False)
                if p.sheet not in wb.sheetnames:
                    raise KeyError(f"Sheet '{p.sheet}' not found.")
                ws = wb[p.sheet]
                ws[p.cell] = p.value
                self._save_wb(p.path, wb)
                return {"written": True, "cell": p.cell, "value": p.value}

        return await asyncio.to_thread(_write)

    async def _action_read_range(self, params: dict[str, Any]) -> dict[str, Any]:
        p = ReadRangeParams.model_validate(params)

        def _read() -> dict[str, Any]:
            with self._get_path_lock(p.path):
                wb = self._get_wb(p.path, data_only=True)
                if p.sheet not in wb.sheetnames:
                    raise KeyError(f"Sheet '{p.sheet}' not found.")
                ws = wb[p.sheet]
                min_row, max_row, min_col, max_col = self._parse_range(p.range, ws)

                rows: list[list[Any]] = []
                for row in ws.iter_rows(
                    min_row=min_row, max_row=max_row,
                    min_col=min_col, max_col=max_col,
                ):
                    rows.append([self._cell_to_value(cell) for cell in row])

                if p.as_dict and p.include_headers and rows:
                    headers = rows[0]
                    data_as_dict = [
                        {str(headers[i]): row[i] for i in range(len(headers))}
                        for row in rows[1:]
                    ]
                    return {
                        "range": p.range,
                        "sheet": p.sheet,
                        "headers": headers,
                        "data": data_as_dict,
                        "row_count": len(data_as_dict),
                    }

                return {
                    "range": p.range,
                    "sheet": p.sheet,
                    "data": rows,
                    "row_count": len(rows),
                    "col_count": len(rows[0]) if rows else 0,
                }

        return await asyncio.to_thread(_read)

    async def _action_write_range(self, params: dict[str, Any]) -> dict[str, Any]:
        p = WriteRangeParams.model_validate(params)

        def _write() -> dict[str, Any]:
            with self._get_path_lock(p.path):
                from openpyxl.utils import column_index_from_string

                wb = self._get_wb(p.path, data_only=False)
                if p.sheet not in wb.sheetnames:
                    raise KeyError(f"Sheet '{p.sheet}' not found.")
                ws = wb[p.sheet]

                # Parse start cell.
                col_str = "".join(c for c in p.start_cell if c.isalpha())
                row_str = "".join(c for c in p.start_cell if c.isdigit())
                start_row = int(row_str)
                start_col = column_index_from_string(col_str)

                cells_written = 0
                for r_idx, row_data in enumerate(p.data):
                    for c_idx, value in enumerate(row_data):
                        ws.cell(row=start_row + r_idx, column=start_col + c_idx, value=value)
                        cells_written += 1

                self._save_wb(p.path, wb)
                return {
                    "written": True,
                    "start_cell": p.start_cell,
                    "rows": len(p.data),
                    "cols": max((len(r) for r in p.data), default=0),
                    "cells_written": cells_written,
                }

        return await asyncio.to_thread(_write)

    async def _action_copy_range(self, params: dict[str, Any]) -> dict[str, Any]:
        p = CopyRangeParams.model_validate(params)

        def _copy() -> dict[str, Any]:
            with self._get_path_lock(p.path):
                import copy as copy_mod

                from openpyxl.utils import column_index_from_string

                wb = self._get_wb(p.path, data_only=False)
                src_ws = wb[p.source_sheet]
                dst_ws = wb[p.dest_sheet]

                min_row, max_row, min_col, max_col = self._parse_range(p.source_range, src_ws)

                col_str = "".join(c for c in p.dest_cell if c.isalpha())
                row_str = "".join(c for c in p.dest_cell if c.isdigit())
                dest_row = int(row_str)
                dest_col = column_index_from_string(col_str)

                cells_copied = 0
                for r_offset, row in enumerate(
                    src_ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col)
                ):
                    for c_offset, src_cell in enumerate(row):
                        dst_cell = dst_ws.cell(row=dest_row + r_offset, column=dest_col + c_offset)
                        dst_cell.value = src_cell.value
                        if not p.copy_values_only and src_cell.has_style:
                            dst_cell.font = copy_mod.copy(src_cell.font)
                            dst_cell.fill = copy_mod.copy(src_cell.fill)
                            dst_cell.border = copy_mod.copy(src_cell.border)
                            dst_cell.alignment = copy_mod.copy(src_cell.alignment)
                            dst_cell.number_format = src_cell.number_format
                        cells_copied += 1

                self._save_wb(p.path, wb)
                return {
                    "copied": True,
                    "cells_copied": cells_copied,
                    "source": f"{p.source_sheet}!{p.source_range}",
                    "destination": f"{p.dest_sheet}!{p.dest_cell}",
                }

        return await asyncio.to_thread(_copy)

    async def _action_insert_rows(self, params: dict[str, Any]) -> dict[str, Any]:
        p = InsertRowsParams.model_validate(params)

        def _insert() -> dict[str, Any]:
            with self._get_path_lock(p.path):
                wb = self._get_wb(p.path, data_only=False)
                ws = wb[p.sheet]
                ws.insert_rows(p.row, p.count)
                self._save_wb(p.path, wb)
                return {"inserted": True, "row": p.row, "count": p.count}

        return await asyncio.to_thread(_insert)

    async def _action_delete_rows(self, params: dict[str, Any]) -> dict[str, Any]:
        p = DeleteRowsParams.model_validate(params)

        def _delete() -> dict[str, Any]:
            with self._get_path_lock(p.path):
                wb = self._get_wb(p.path, data_only=False)
                ws = wb[p.sheet]
                ws.delete_rows(p.row, p.count)
                self._save_wb(p.path, wb)
                return {"deleted": True, "row": p.row, "count": p.count}

        return await asyncio.to_thread(_delete)

    async def _action_insert_columns(self, params: dict[str, Any]) -> dict[str, Any]:
        p = InsertColumnsParams.model_validate(params)

        def _insert() -> dict[str, Any]:
            with self._get_path_lock(p.path):
                wb = self._get_wb(p.path, data_only=False)
                ws = wb[p.sheet]
                ws.insert_cols(p.column, p.count)
                self._save_wb(p.path, wb)
                return {"inserted": True, "column": p.column, "count": p.count}

        return await asyncio.to_thread(_insert)

    async def _action_delete_columns(self, params: dict[str, Any]) -> dict[str, Any]:
        p = DeleteColumnsParams.model_validate(params)

        def _delete() -> dict[str, Any]:
            with self._get_path_lock(p.path):
                wb = self._get_wb(p.path, data_only=False)
                ws = wb[p.sheet]
                ws.delete_cols(p.column, p.count)
                self._save_wb(p.path, wb)
                return {"deleted": True, "column": p.column, "count": p.count}

        return await asyncio.to_thread(_delete)

    async def _action_merge_cells(self, params: dict[str, Any]) -> dict[str, Any]:
        p = MergeCellsParams.model_validate(params)

        def _merge() -> dict[str, Any]:
            with self._get_path_lock(p.path):
                wb = self._get_wb(p.path, data_only=False)
                ws = wb[p.sheet]
                ws.merge_cells(p.range)
                self._save_wb(p.path, wb)
                return {"merged": True, "range": p.range}

        return await asyncio.to_thread(_merge)

    async def _action_unmerge_cells(self, params: dict[str, Any]) -> dict[str, Any]:
        p = UnmergeCellsParams.model_validate(params)

        def _unmerge() -> dict[str, Any]:
            with self._get_path_lock(p.path):
                wb = self._get_wb(p.path, data_only=False)
                ws = wb[p.sheet]
                ws.unmerge_cells(p.range)
                self._save_wb(p.path, wb)
                return {"unmerged": True, "range": p.range}

        return await asyncio.to_thread(_unmerge)

    async def _action_freeze_panes(self, params: dict[str, Any]) -> dict[str, Any]:
        p = FreezePanesParams.model_validate(params)

        def _freeze() -> dict[str, Any]:
            with self._get_path_lock(p.path):
                wb = self._get_wb(p.path, data_only=False)
                if p.sheet not in wb.sheetnames:
                    raise KeyError(f"Sheet '{p.sheet}' not found.")
                ws = wb[p.sheet]
                ws.freeze_panes = p.cell  # None unfreeze
                self._save_wb(p.path, wb)
                return {"freeze_panes": p.cell, "sheet": p.sheet}

        return await asyncio.to_thread(_freeze)

    async def _action_set_column_width(self, params: dict[str, Any]) -> dict[str, Any]:
        p = SetColumnWidthParams.model_validate(params)

        def _set_width() -> dict[str, Any]:
            with self._get_path_lock(p.path):
                wb = self._get_wb(p.path, data_only=False)
                if p.sheet not in wb.sheetnames:
                    raise KeyError(f"Sheet '{p.sheet}' not found.")
                ws = wb[p.sheet]

                # Resolve column letters — support range like "A:D"
                from openpyxl.utils import column_index_from_string, get_column_letter

                if ":" in p.column:
                    col_start, col_end = p.column.split(":", 1)
                    col_indices = range(
                        column_index_from_string(col_start.strip()),
                        column_index_from_string(col_end.strip()) + 1,
                    )
                else:
                    col_indices = [column_index_from_string(p.column.strip())]

                updated: list[str] = []
                for col_idx in col_indices:
                    col_letter = get_column_letter(col_idx)
                    if p.auto_fit or p.width is None:
                        # Auto-fit: iterate all rows to find longest content.
                        max_len = 0
                        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
                            for cell in row:
                                if cell.value is not None:
                                    max_len = max(max_len, len(str(cell.value)))
                        ws.column_dimensions[col_letter].width = max(max_len + 2, 8)
                    else:
                        ws.column_dimensions[col_letter].width = p.width
                    updated.append(col_letter)

                self._save_wb(p.path, wb)
                return {"updated_columns": updated, "sheet": p.sheet}

        return await asyncio.to_thread(_set_width)

    async def _action_set_row_height(self, params: dict[str, Any]) -> dict[str, Any]:
        p = SetRowHeightParams.model_validate(params)

        def _set_height() -> dict[str, Any]:
            with self._get_path_lock(p.path):
                wb = self._get_wb(p.path, data_only=False)
                if p.sheet not in wb.sheetnames:
                    raise KeyError(f"Sheet '{p.sheet}' not found.")
                ws = wb[p.sheet]
                ws.row_dimensions[p.row].height = p.height
                self._save_wb(p.path, wb)
                return {"row": p.row, "height": p.height, "sheet": p.sheet}

        return await asyncio.to_thread(_set_height)

    async def _action_find_replace(self, params: dict[str, Any]) -> dict[str, Any]:
        p = FindReplaceParams.model_validate(params)

        def _find_replace() -> dict[str, Any]:
            with self._get_path_lock(p.path):
                wb = self._get_wb(p.path, data_only=False)

                sheets_to_scan = [wb[p.sheet]] if p.sheet else [wb[s] for s in wb.sheetnames]

                replacements = 0
                find_val = p.find if p.case_sensitive else p.find.lower()

                for ws in sheets_to_scan:
                    for row in ws.iter_rows():
                        for cell in row:
                            if cell.value is None:
                                continue
                            cell_str = str(cell.value)
                            compare_str = cell_str if p.case_sensitive else cell_str.lower()

                            if p.whole_cell:
                                if compare_str == find_val:
                                    cell.value = p.replace
                                    replacements += 1
                            else:
                                if find_val in compare_str:
                                    if p.case_sensitive:
                                        cell.value = cell_str.replace(p.find, p.replace)
                                    else:
                                        import re
                                        cell.value = re.sub(
                                            re.escape(p.find), p.replace, cell_str, flags=re.IGNORECASE
                                        )
                                    replacements += 1

                self._save_wb(p.path, wb)
                return {"replacements": replacements, "find": p.find, "replace": p.replace}

        return await asyncio.to_thread(_find_replace)

    async def _action_remove_duplicates(self, params: dict[str, Any]) -> dict[str, Any]:
        p = RemoveDuplicatesParams.model_validate(params)

        def _remove_dups() -> dict[str, Any]:
            with self._get_path_lock(p.path):
                wb = self._get_wb(p.path, data_only=False)
                if p.sheet not in wb.sheetnames:
                    raise KeyError(f"Sheet '{p.sheet}' not found.")
                ws = wb[p.sheet]
                min_row, max_row, min_col, max_col = self._parse_range(p.range, ws)

                # Read all rows in the range.
                rows_data: list[list[Any]] = []
                for row in ws.iter_rows(
                    min_row=min_row, max_row=max_row,
                    min_col=min_col, max_col=max_col,
                ):
                    rows_data.append([self._cell_to_value(c) for c in row])

                # Determine key columns (0-based within the extracted range).
                key_cols = p.columns if p.columns is not None else list(range(len(rows_data[0]) if rows_data else 0))

                seen: set[tuple[Any, ...]] = set()
                unique_rows: list[list[Any]] = []
                duplicates_removed = 0

                iteration = rows_data if p.keep == "first" else reversed(rows_data)
                for row in iteration:
                    key = tuple(row[i] for i in key_cols if i < len(row))
                    if key not in seen:
                        seen.add(key)
                        unique_rows.append(row)
                    else:
                        duplicates_removed += 1

                if p.keep == "last":
                    unique_rows = list(reversed(unique_rows))

                # Write back unique rows, clear remaining cells.
                for r_offset, row_data in enumerate(unique_rows):
                    for c_offset, value in enumerate(row_data):
                        ws.cell(row=min_row + r_offset, column=min_col + c_offset, value=value)

                # Clear rows beyond unique count.
                for r_offset in range(len(unique_rows), len(rows_data)):
                    for c_offset in range(max_col - min_col + 1):
                        ws.cell(row=min_row + r_offset, column=min_col + c_offset, value=None)

                self._save_wb(p.path, wb)
                return {
                    "duplicates_removed": duplicates_removed,
                    "unique_rows": len(unique_rows),
                    "sheet": p.sheet,
                }

        return await asyncio.to_thread(_remove_dups)

    # ------------------------------------------------------------------
    # Formulas & logic
    # ------------------------------------------------------------------

    async def _action_apply_formula(self, params: dict[str, Any]) -> dict[str, Any]:
        p = ApplyFormulaParams.model_validate(params)

        def _apply() -> dict[str, Any]:
            with self._get_path_lock(p.path):
                wb = self._get_wb(p.path, data_only=False)
                if p.sheet not in wb.sheetnames:
                    raise KeyError(f"Sheet '{p.sheet}' not found.")
                ws = wb[p.sheet]
                ws[p.cell] = p.formula
                self._save_wb(p.path, wb)
                return {"applied": True, "cell": p.cell, "formula": p.formula}

        return await asyncio.to_thread(_apply)

    async def _action_add_named_range(self, params: dict[str, Any]) -> dict[str, Any]:
        p = AddNamedRangeParams.model_validate(params)

        def _add() -> dict[str, Any]:
            from openpyxl.workbook.defined_name import DefinedName

            wb = self._get_wb(p.path, data_only=False)
            if p.sheet not in wb.sheetnames:
                raise KeyError(f"Sheet '{p.sheet}' not found.")

            attr_text = f"'{p.sheet}'!{p.range}"

            if p.scope == "workbook":
                dn = DefinedName(name=p.name, attr_text=attr_text)
                wb.defined_names.add(dn)
            else:
                # Sheet-scoped: assign localSheetId
                sheet_id = wb.sheetnames.index(p.sheet)
                dn = DefinedName(name=p.name, attr_text=attr_text, localSheetId=sheet_id)
                wb.defined_names.add(dn)

            self._save_wb(p.path, wb)
            return {"added": True, "name": p.name, "range": attr_text, "scope": p.scope}

        return await asyncio.to_thread(_add)

    # ------------------------------------------------------------------
    # Formatting
    # ------------------------------------------------------------------

    async def _action_format_range(self, params: dict[str, Any]) -> dict[str, Any]:
        p = FormatRangeParams.model_validate(params)

        def _format() -> dict[str, Any]:
            from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

            wb = self._get_wb(p.path, data_only=False)
            if p.sheet not in wb.sheetnames:
                raise KeyError(f"Sheet '{p.sheet}' not found.")
            ws = wb[p.sheet]
            min_row, max_row, min_col, max_col = self._parse_range(p.range, ws)

            # Build font kwargs.
            font_kwargs: dict[str, Any] = {}
            if p.bold is not None:
                font_kwargs["bold"] = p.bold
            if p.italic is not None:
                font_kwargs["italic"] = p.italic
            if p.underline is not None:
                font_kwargs["underline"] = "single" if p.underline else None
            if p.strikethrough is not None:
                font_kwargs["strike"] = p.strikethrough
            if p.font_name is not None:
                font_kwargs["name"] = p.font_name
            if p.font_size is not None:
                font_kwargs["size"] = p.font_size
            if p.font_color is not None:
                font_kwargs["color"] = p.font_color.lstrip("#")

            # Build fill.
            fill = None
            if p.fill_color is not None:
                fill_type = p.fill_type or "solid"
                if fill_type == "none":
                    fill = PatternFill(fill_type=None)
                else:
                    fill = PatternFill(
                        fill_type="solid",
                        fgColor=p.fill_color.lstrip("#"),
                    )

            # Build border.
            border = None
            if p.border_style is not None and p.border_style != "none":
                border_color = (p.border_color or "000000").lstrip("#")
                side = Side(style=p.border_style, color=border_color)
                sides = p.border_sides or ["all"]

                if "all" in sides or "outline" in sides:
                    border = Border(left=side, right=side, top=side, bottom=side)
                else:
                    border = Border(
                        left=side if "left" in sides else None,
                        right=side if "right" in sides else None,
                        top=side if "top" in sides else None,
                        bottom=side if "bottom" in sides else None,
                    )

            # Build alignment.
            alignment = None
            if any(v is not None for v in [p.alignment_horizontal, p.alignment_vertical, p.wrap_text]):
                alignment = Alignment(
                    horizontal=p.alignment_horizontal,
                    vertical=p.alignment_vertical,
                    wrap_text=p.wrap_text,
                )

            cells_formatted = 0
            for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
                for cell in row:
                    if font_kwargs:
                        import copy as copy_mod
                        existing = copy_mod.copy(cell.font)
                        for k, v in font_kwargs.items():
                            setattr(existing, k, v)
                        cell.font = existing
                    if fill is not None:
                        cell.fill = fill
                    if border is not None:
                        cell.border = border
                    if alignment is not None:
                        cell.alignment = alignment
                    if p.number_format is not None:
                        cell.number_format = p.number_format
                    cells_formatted += 1

            self._save_wb(p.path, wb)
            return {"formatted": True, "cells_formatted": cells_formatted, "range": p.range}

        return await asyncio.to_thread(_format)

    async def _action_apply_conditional_format(self, params: dict[str, Any]) -> dict[str, Any]:
        p = ApplyConditionalFormatParams.model_validate(params)

        def _apply_cf() -> dict[str, Any]:
            from openpyxl.formatting.rule import (
                ColorScaleRule,
                DataBarRule,
                FormulaRule,
                IconSetRule,
            )
            from openpyxl.styles import Font, PatternFill

            wb = self._get_wb(p.path, data_only=False)
            if p.sheet not in wb.sheetnames:
                raise KeyError(f"Sheet '{p.sheet}' not found.")
            ws = wb[p.sheet]

            rule: Any = None

            if p.format_type == "color_scale":
                if p.mid_color:
                    rule = ColorScaleRule(
                        start_type="min", start_color=p.min_color.lstrip("#"),
                        mid_type="percentile", mid_value=50, mid_color=p.mid_color.lstrip("#"),
                        end_type="max", end_color=p.max_color.lstrip("#"),
                    )
                else:
                    rule = ColorScaleRule(
                        start_type="min", start_color=p.min_color.lstrip("#"),
                        end_type="max", end_color=p.max_color.lstrip("#"),
                    )

            elif p.format_type == "data_bar":
                rule = DataBarRule(
                    start_type="min", start_value=0,
                    end_type="max", end_value=100,
                    color=p.max_color.lstrip("#"),
                )

            elif p.format_type == "icon_set":
                rule = IconSetRule(icon_style="3Arrows", type="percent", values=[0, 33, 67])

            elif p.format_type == "cell_is":
                from openpyxl.formatting.rule import CellIsRule

                # Build the differential formatting object.
                dxf_font = Font(bold=p.bold, color=p.font_color.lstrip("#") if p.font_color else None) if (p.bold or p.font_color) else None
                dxf_fill = PatternFill(fill_type="solid", fgColor=p.fill_color.lstrip("#")) if p.fill_color else None

                values: list[Any] = [p.value] if p.value is not None else []
                if p.value2 is not None:
                    values.append(p.value2)

                rule = CellIsRule(
                    operator=p.operator or "greaterThan",
                    formula=values,
                    font=dxf_font,
                    fill=dxf_fill,
                )

            elif p.format_type == "formula":
                dxf_font = Font(bold=p.bold, color=p.font_color.lstrip("#") if p.font_color else None) if (p.bold or p.font_color) else None
                dxf_fill = PatternFill(fill_type="solid", fgColor=p.fill_color.lstrip("#")) if p.fill_color else None
                rule = FormulaRule(
                    formula=[p.formula or "TRUE"],
                    font=dxf_font,
                    fill=dxf_fill,
                )

            if rule is not None:
                ws.conditional_formatting.add(p.range, rule)

            self._save_wb(p.path, wb)
            return {"applied": True, "format_type": p.format_type, "range": p.range}

        return await asyncio.to_thread(_apply_cf)

    async def _action_add_data_validation(self, params: dict[str, Any]) -> dict[str, Any]:
        p = AddDataValidationParams.model_validate(params)

        def _add_dv() -> dict[str, Any]:
            from openpyxl.worksheet.datavalidation import DataValidation

            wb = self._get_wb(p.path, data_only=False)
            if p.sheet not in wb.sheetnames:
                raise KeyError(f"Sheet '{p.sheet}' not found.")
            ws = wb[p.sheet]

            dv = DataValidation(
                type=p.validation_type,
                operator=p.operator,
                formula1=p.formula1,
                formula2=p.formula2,
                allow_blank=p.allow_blank,
                showDropDown=not p.show_dropdown,
                showInputMessage=p.show_input_message,
                showErrorMessage=p.show_error_alert,
                promptTitle=p.input_title,
                prompt=p.input_message,
                errorTitle=p.error_title,
                error=p.error_message,
                errorStyle=p.error_style,
            )
            dv.sqref = p.range
            ws.add_data_validation(dv)

            self._save_wb(p.path, wb)
            return {"added": True, "validation_type": p.validation_type, "range": p.range}

        return await asyncio.to_thread(_add_dv)

    async def _action_add_autofilter(self, params: dict[str, Any]) -> dict[str, Any]:
        p = AddAutoFilterParams.model_validate(params)

        def _add_af() -> dict[str, Any]:
            wb = self._get_wb(p.path, data_only=False)
            if p.sheet not in wb.sheetnames:
                raise KeyError(f"Sheet '{p.sheet}' not found.")
            ws = wb[p.sheet]
            ws.auto_filter.ref = p.range if p.range else ws.dimensions
            self._save_wb(p.path, wb)
            return {"added": True, "range": ws.auto_filter.ref, "sheet": p.sheet}

        return await asyncio.to_thread(_add_af)

    # ------------------------------------------------------------------
    # Charts
    # ------------------------------------------------------------------

    async def _action_create_chart(self, params: dict[str, Any]) -> dict[str, Any]:
        p = CreateChartParams.model_validate(params)

        def _create_chart() -> dict[str, Any]:
            from openpyxl.chart import (
                AreaChart,
                BarChart,
                DoughnutChart,
                LineChart,
                PieChart,
                RadarChart,
                ScatterChart,
            )
            from openpyxl.chart import Reference
            from openpyxl.utils import column_index_from_string

            wb = self._get_wb(p.path, data_only=False)
            if p.sheet not in wb.sheetnames:
                raise KeyError(f"Sheet '{p.sheet}' not found.")
            ws = wb[p.sheet]

            # Map chart_type string to class.
            chart_map: dict[str, Any] = {
                "bar": BarChart,
                "col": BarChart,
                "line": LineChart,
                "pie": PieChart,
                "doughnut": DoughnutChart,
                "scatter": ScatterChart,
                "area": AreaChart,
                "radar": RadarChart,
            }

            chart_cls = chart_map.get(p.chart_type, BarChart)
            chart = chart_cls()

            if p.chart_type == "col":
                chart.type = "col"
            elif p.chart_type == "bar":
                chart.type = "bar"

            chart.style = p.style
            if p.title:
                chart.title = p.title

            # Axis titles where applicable.
            if hasattr(chart, "x_axis") and p.x_axis_title:
                chart.x_axis.title = p.x_axis_title
            if hasattr(chart, "y_axis") and p.y_axis_title:
                chart.y_axis.title = p.y_axis_title

            if hasattr(chart, "grouping"):
                try:
                    chart.grouping = p.grouping
                except Exception:
                    # Some chart types (e.g. LineChart) don't support "clustered".
                    pass

            chart.legend = None if not p.has_legend else chart.legend

            # Parse data range into a Reference.
            min_row, max_row, min_col, max_col = self._parse_range(p.data_range, ws)
            data_ref = Reference(
                ws,
                min_row=min_row,
                max_row=max_row,
                min_col=min_col,
                max_col=max_col,
            )

            if isinstance(chart, ScatterChart):
                from openpyxl.chart import Series

                # For scatter: x = first col, y = remaining cols.
                x_ref = Reference(ws, min_row=min_row, max_row=max_row, min_col=min_col)
                for col_idx in range(min_col + 1, max_col + 1):
                    y_ref = Reference(ws, min_row=min_row, max_row=max_row, min_col=col_idx)
                    series = Series(y_ref, x_ref)
                    chart.series.append(series)
            else:
                titles_from = 1 if p.series_labels else None
                chart.add_data(data_ref, titles_from_data=p.series_labels)
                if p.series_labels and (max_row > min_row or max_col > min_col):
                    try:
                        cats_ref = Reference(ws, min_row=min_row + 1, max_row=max_row, min_col=min_col)
                        chart.set_categories(cats_ref)
                    except Exception:
                        pass

            # Set dimensions.
            from openpyxl.utils.units import cm_to_EMU

            chart.width = p.width
            chart.height = p.height

            ws.add_chart(chart, p.position)
            self._save_wb(p.path, wb)
            return {
                "created": True,
                "chart_type": p.chart_type,
                "position": p.position,
                "data_range": p.data_range,
            }

        return await asyncio.to_thread(_create_chart)

    # ------------------------------------------------------------------
    # Images
    # ------------------------------------------------------------------

    async def _action_insert_image(self, params: dict[str, Any]) -> dict[str, Any]:
        p = InsertImageParams.model_validate(params)

        def _insert_img() -> dict[str, Any]:
            from openpyxl.drawing.image import Image

            wb = self._get_wb(p.path, data_only=False)
            if p.sheet not in wb.sheetnames:
                raise KeyError(f"Sheet '{p.sheet}' not found.")
            ws = wb[p.sheet]

            img = Image(p.image_path)
            if p.width is not None:
                img.width = p.width
            if p.height is not None:
                img.height = p.height

            ws.add_image(img, p.cell)
            self._save_wb(p.path, wb)
            return {
                "inserted": True,
                "image_path": p.image_path,
                "cell": p.cell,
                "sheet": p.sheet,
            }

        return await asyncio.to_thread(_insert_img)

    # ------------------------------------------------------------------
    # Comments
    # ------------------------------------------------------------------

    async def _action_add_comment(self, params: dict[str, Any]) -> dict[str, Any]:
        p = AddCommentParams.model_validate(params)

        def _add_comment() -> dict[str, Any]:
            from openpyxl.comments import Comment

            wb = self._get_wb(p.path, data_only=False)
            if p.sheet not in wb.sheetnames:
                raise KeyError(f"Sheet '{p.sheet}' not found.")
            ws = wb[p.sheet]
            comment = Comment(p.text, p.author)
            comment.width = p.width
            comment.height = p.height
            ws[p.cell].comment = comment
            self._save_wb(p.path, wb)
            return {"added": True, "cell": p.cell, "author": p.author}

        return await asyncio.to_thread(_add_comment)

    async def _action_delete_comment(self, params: dict[str, Any]) -> dict[str, Any]:
        p = DeleteCommentParams.model_validate(params)

        def _del_comment() -> dict[str, Any]:
            wb = self._get_wb(p.path, data_only=False)
            if p.sheet not in wb.sheetnames:
                raise KeyError(f"Sheet '{p.sheet}' not found.")
            ws = wb[p.sheet]
            ws[p.cell].comment = None
            self._save_wb(p.path, wb)
            return {"deleted": True, "cell": p.cell}

        return await asyncio.to_thread(_del_comment)

    # ------------------------------------------------------------------
    # Export
    # ------------------------------------------------------------------

    async def _action_export_to_csv(self, params: dict[str, Any]) -> dict[str, Any]:
        p = ExportToCsvParams.model_validate(params)

        def _export_csv() -> dict[str, Any]:
            wb = self._get_wb(p.path, data_only=True)
            if p.sheet not in wb.sheetnames:
                raise KeyError(f"Sheet '{p.sheet}' not found.")
            ws = wb[p.sheet]

            output_path = Path(p.output_path)
            output_path.parent.mkdir(parents=True, exist_ok=True)

            rows_written = 0
            with output_path.open("w", newline="", encoding=p.encoding) as f:
                writer = csv.writer(f, delimiter=p.delimiter)
                for r_idx, row in enumerate(ws.iter_rows(values_only=True)):
                    if r_idx == 0 and not p.include_header:
                        continue
                    writer.writerow(row)
                    rows_written += 1

            return {
                "exported": True,
                "output_path": str(output_path),
                "rows_written": rows_written,
                "sheet": p.sheet,
            }

        return await asyncio.to_thread(_export_csv)

    async def _action_export_to_pdf(self, params: dict[str, Any]) -> dict[str, Any]:
        p = ExportToPdfParams.model_validate(params)

        def _export_pdf() -> dict[str, Any]:
            src_path = str(Path(p.path).resolve())
            output_dir = str(Path(p.output_path).parent.resolve())
            Path(output_dir).mkdir(parents=True, exist_ok=True)

            if not p.use_libreoffice:
                raise RuntimeError("Only LibreOffice-based PDF export is supported. Set use_libreoffice=True.")

            cmd = [
                "libreoffice",
                "--headless",
                "--convert-to", "pdf",
                "--outdir", output_dir,
                src_path,
            ]
            result = subprocess.run(cmd, capture_output=True, text=True, timeout=120)
            if result.returncode != 0:
                raise RuntimeError(
                    f"LibreOffice conversion failed (exit {result.returncode}): {result.stderr.strip()}"
                )

            # LibreOffice creates a file with same name but .pdf extension.
            generated_pdf = Path(output_dir) / (Path(src_path).stem + ".pdf")
            # Rename to requested output path if different.
            desired = Path(p.output_path).resolve()
            if generated_pdf.resolve() != desired:
                generated_pdf.rename(desired)

            return {
                "exported": True,
                "output_path": str(desired),
                "source": src_path,
            }

        return await asyncio.to_thread(_export_pdf)

    # ------------------------------------------------------------------
    # Manifest
    # ------------------------------------------------------------------

    def get_manifest(self) -> ModuleManifest:
        return ModuleManifest(
            module_id=self.MODULE_ID,
            version=self.VERSION,
            description="Full-featured Excel spreadsheet automation using openpyxl.",
            platforms=["all"],
            tags=["excel", "spreadsheet", "office", "xlsx"],
            dependencies=["openpyxl>=3.1"],
            declared_permissions=["filesystem_read", "filesystem_write"],
            actions=[
                # ── Workbook lifecycle ────────────────────────────
                ActionSpec(
                    name="create_workbook",
                    description="Create a new blank Excel workbook (.xlsx) with an initial sheet.",
                    params=[
                        ParamSpec("path", "string", "Path for the new .xlsx file."),
                        ParamSpec("sheet_name", "string", "Name of the initial worksheet.", required=False, default="Sheet1"),
                    ],
                    returns="object",
                    returns_description='{"path": str, "created": true, "sheet_names": list[str]}',
                    permission_required="local_worker",
                    tags=["workbook", "create"],
                    examples=[
                        {"description": "Create an empty workbook", "params": {"path": "/data/new.xlsx"}},
                    ],
                ),
                ActionSpec(
                    name="open_workbook",
                    description="Open an Excel workbook (.xlsx/.xlsm) and cache it for subsequent operations.",
                    params=[
                        ParamSpec("path", "string", "Path to the .xlsx / .xlsm file."),
                        ParamSpec("read_only", "boolean", "Open in read-only mode.", required=False, default=False),
                        ParamSpec("data_only", "boolean", "Return cached formula results instead of formulas.", required=False, default=True),
                        ParamSpec("keep_vba", "boolean", "Preserve VBA macros (.xlsm).", required=False, default=False),
                    ],
                    returns="object",
                    returns_description='{"path": str, "sheet_names": list[str], "active_sheet": str}',
                    permission_required="local_worker",
                    tags=["workbook", "open"],
                    examples=[
                        {"description": "Open a workbook", "params": {"path": "/data/sales.xlsx"}},
                    ],
                ),
                ActionSpec(
                    name="close_workbook",
                    description="Close a cached workbook, optionally saving it first.",
                    params=[
                        ParamSpec("path", "string", "Path to the workbook."),
                        ParamSpec("save", "boolean", "Save before closing.", required=False, default=False),
                    ],
                    returns="object",
                    returns_description='{"closed": true, "path": str}',
                    permission_required="local_worker",
                    tags=["workbook", "close"],
                ),
                ActionSpec(
                    name="save_workbook",
                    description="Save a workbook to disk (optionally to a new path).",
                    params=[
                        ParamSpec("path", "string", "Path to the workbook."),
                        ParamSpec("output_path", "string", "Save As path. Overwrites original if not set.", required=False),
                    ],
                    returns="object",
                    returns_description='{"saved": true, "path": str}',
                    permission_required="local_worker",
                    tags=["workbook", "save"],
                ),
                ActionSpec(
                    name="get_workbook_info",
                    description="Retrieve metadata about a workbook: sheets, named ranges, defined names.",
                    params=[
                        ParamSpec("path", "string", "Path to the workbook."),
                        ParamSpec("include_named_ranges", "boolean", "Include named ranges.", required=False, default=True),
                        ParamSpec("include_defined_names", "boolean", "Include defined names.", required=False, default=True),
                    ],
                    returns="object",
                    returns_description='{"path": str, "sheet_names": list, "named_ranges": list, ...}',
                    permission_required="local_worker",
                    tags=["workbook", "info"],
                ),
                # ── Sheet management ──────────────────────────────
                ActionSpec(
                    name="list_sheets",
                    description="List all worksheet names in a workbook.",
                    params=[
                        ParamSpec("path", "string", "Path to the workbook."),
                    ],
                    returns="object",
                    returns_description='{"sheets": list[str]}',
                    permission_required="local_worker",
                    tags=["sheet", "list"],
                ),
                ActionSpec(
                    name="create_sheet",
                    description="Add a new worksheet to a workbook.",
                    params=[
                        ParamSpec("path", "string", "Path to the workbook."),
                        ParamSpec("name", "string", "Name for the new sheet."),
                        ParamSpec("position", "integer", "Insert position (0-indexed). Appended if not set.", required=False),
                    ],
                    returns="object",
                    returns_description='{"created": true, "name": str, "sheet_names": list}',
                    permission_required="local_worker",
                    tags=["sheet", "create"],
                ),
                ActionSpec(
                    name="delete_sheet",
                    description="Delete a worksheet from a workbook.",
                    params=[
                        ParamSpec("path", "string", "Path to the workbook."),
                        ParamSpec("name", "string", "Name of the sheet to delete."),
                    ],
                    returns="object",
                    returns_description='{"deleted": true, "sheet_names": list}',
                    permission_required="local_worker",
                    tags=["sheet", "delete"],
                ),
                ActionSpec(
                    name="rename_sheet",
                    description="Rename a worksheet.",
                    params=[
                        ParamSpec("path", "string", "Path to the workbook."),
                        ParamSpec("old_name", "string", "Current sheet name."),
                        ParamSpec("new_name", "string", "New sheet name."),
                    ],
                    returns="object",
                    returns_description='{"renamed": true, "old_name": str, "new_name": str}',
                    permission_required="local_worker",
                    tags=["sheet", "rename"],
                ),
                ActionSpec(
                    name="copy_sheet",
                    description="Copy a worksheet within the same workbook.",
                    params=[
                        ParamSpec("path", "string", "Path to the workbook."),
                        ParamSpec("source_sheet", "string", "Name of the sheet to copy."),
                        ParamSpec("new_name", "string", "Name for the copy."),
                        ParamSpec("position", "integer", "Insert position (0-indexed).", required=False),
                    ],
                    returns="object",
                    returns_description='{"copied": true, "new_name": str, "sheet_names": list}',
                    permission_required="local_worker",
                    tags=["sheet", "copy"],
                ),
                ActionSpec(
                    name="get_sheet_info",
                    description="Get information about a worksheet: dimensions, merged cells, etc.",
                    params=[
                        ParamSpec("path", "string", "Path to the workbook."),
                        ParamSpec("sheet", "string", "Worksheet name."),
                        ParamSpec("include_dimensions", "boolean", "Include dimension info.", required=False, default=True),
                        ParamSpec("include_merged_cells", "boolean", "Include merged cell ranges.", required=False, default=True),
                    ],
                    returns="object",
                    returns_description='{"sheet": str, "dimensions": str, "merged_cells": list, ...}',
                    permission_required="local_worker",
                    tags=["sheet", "info"],
                ),
                ActionSpec(
                    name="protect_sheet",
                    description="Protect a worksheet with an optional password.",
                    params=[
                        ParamSpec("path", "string", "Path to the workbook."),
                        ParamSpec("sheet", "string", "Worksheet name."),
                        ParamSpec("password", "string", "Protection password.", required=False),
                        ParamSpec("allow_select_locked", "boolean", "Allow selecting locked cells.", required=False, default=True),
                        ParamSpec("allow_select_unlocked", "boolean", "Allow selecting unlocked cells.", required=False, default=True),
                        ParamSpec("allow_format_cells", "boolean", "Allow formatting cells.", required=False, default=False),
                        ParamSpec("allow_sort", "boolean", "Allow sorting.", required=False, default=False),
                        ParamSpec("allow_auto_filter", "boolean", "Allow auto-filter.", required=False, default=False),
                        ParamSpec("allow_insert_rows", "boolean", "Allow inserting rows.", required=False, default=False),
                        ParamSpec("allow_delete_rows", "boolean", "Allow deleting rows.", required=False, default=False),
                    ],
                    returns="object",
                    returns_description='{"protected": true, "sheet": str}',
                    permission_required="local_worker",
                    tags=["sheet", "protect"],
                ),
                ActionSpec(
                    name="unprotect_sheet",
                    description="Remove protection from a worksheet.",
                    params=[
                        ParamSpec("path", "string", "Path to the workbook."),
                        ParamSpec("sheet", "string", "Worksheet name."),
                        ParamSpec("password", "string", "Protection password.", required=False),
                    ],
                    returns="object",
                    returns_description='{"unprotected": true, "sheet": str}',
                    permission_required="local_worker",
                    tags=["sheet", "unprotect"],
                ),
                ActionSpec(
                    name="set_page_setup",
                    description="Configure page setup (orientation, margins, paper size, print area) for a worksheet.",
                    params=[
                        ParamSpec("path", "string", "Path to the workbook."),
                        ParamSpec("sheet", "string", "Worksheet name."),
                        ParamSpec("orientation", "string", "Page orientation: portrait or landscape.", required=False, default="portrait"),
                        ParamSpec("paper_size", "integer", "Paper size code (9=A4, 1=Letter).", required=False, default=9),
                        ParamSpec("fit_to_page", "boolean", "Fit content to page.", required=False, default=False),
                        ParamSpec("top_margin", "number", "Top margin in inches.", required=False, default=1.0),
                        ParamSpec("bottom_margin", "number", "Bottom margin in inches.", required=False, default=1.0),
                        ParamSpec("left_margin", "number", "Left margin in inches.", required=False, default=0.75),
                        ParamSpec("right_margin", "number", "Right margin in inches.", required=False, default=0.75),
                        ParamSpec("print_area", "string", "Print area range, e.g. 'A1:H50'.", required=False),
                    ],
                    returns="object",
                    returns_description='{"configured": true, "sheet": str}',
                    permission_required="local_worker",
                    tags=["sheet", "page", "print"],
                ),
                # ── Cell & range operations ───────────────────────
                ActionSpec(
                    name="read_cell",
                    description="Read the value of a single cell.",
                    params=[
                        ParamSpec("path", "string", "Path to the workbook."),
                        ParamSpec("sheet", "string", "Worksheet name."),
                        ParamSpec("cell", "string", "Cell address, e.g. 'B3'.", example="B3"),
                    ],
                    returns="object",
                    returns_description='{"cell": str, "value": Any, "data_type": str}',
                    permission_required="local_worker",
                    tags=["read", "cell"],
                ),
                ActionSpec(
                    name="write_cell",
                    description="Write a value to a single cell.",
                    params=[
                        ParamSpec("path", "string", "Path to the workbook."),
                        ParamSpec("sheet", "string", "Worksheet name."),
                        ParamSpec("cell", "string", "Cell address, e.g. 'B3'.", example="B3"),
                        ParamSpec("value", "string", "Value to write (string, number, boolean, or null)."),
                    ],
                    returns="object",
                    returns_description='{"written": true, "cell": str, "value": Any}',
                    permission_required="local_worker",
                    tags=["write", "cell"],
                ),
                ActionSpec(
                    name="read_range",
                    description="Read a rectangular range of cells from a worksheet. Use range='auto' for the entire used area.",
                    params=[
                        ParamSpec("path", "string", "Path to the workbook."),
                        ParamSpec("sheet", "string", "Worksheet name."),
                        ParamSpec("range", "string", "Cell range (e.g. 'A1:Z100') or 'auto'.", example="A1:D20"),
                        ParamSpec("include_headers", "boolean", "Treat first row as column headers.", required=False, default=True),
                        ParamSpec("as_dict", "boolean", "Return list[dict] keyed by header.", required=False, default=False),
                    ],
                    returns="object",
                    returns_description='{"data": list[list], "row_count": int, "col_count": int}',
                    permission_required="local_worker",
                    tags=["read", "range", "data"],
                    examples=[
                        {"description": "Read a data table with headers", "params": {"path": "/data/sales.xlsx", "sheet": "Sheet1", "range": "A1:D100", "as_dict": True}},
                    ],
                ),
                ActionSpec(
                    name="write_range",
                    description="Write a 2-D list of values starting at a given cell.",
                    params=[
                        ParamSpec("path", "string", "Path to the workbook."),
                        ParamSpec("sheet", "string", "Worksheet name."),
                        ParamSpec("start_cell", "string", "Top-left cell of the range.", example="A1"),
                        ParamSpec("data", "array", "Row-major list of rows to write."),
                    ],
                    returns="object",
                    returns_description='{"written": true, "rows": int, "cells_written": int}',
                    permission_required="local_worker",
                    tags=["write", "range", "data"],
                ),
                ActionSpec(
                    name="copy_range",
                    description="Copy a range of cells to another location (same or different sheet).",
                    params=[
                        ParamSpec("path", "string", "Path to the workbook."),
                        ParamSpec("source_sheet", "string", "Source worksheet name."),
                        ParamSpec("source_range", "string", "Source cell range.", example="A1:D10"),
                        ParamSpec("dest_sheet", "string", "Destination worksheet name."),
                        ParamSpec("dest_cell", "string", "Top-left destination cell.", example="F1"),
                        ParamSpec("copy_values_only", "boolean", "Copy only values, not formatting.", required=False, default=False),
                    ],
                    returns="object",
                    returns_description='{"copied": true, "cells_copied": int}',
                    permission_required="local_worker",
                    tags=["copy", "range"],
                ),
                # ── Row & column operations ───────────────────────
                ActionSpec(
                    name="insert_rows",
                    description="Insert empty rows before a given row.",
                    params=[
                        ParamSpec("path", "string", "Path to the workbook."),
                        ParamSpec("sheet", "string", "Worksheet name."),
                        ParamSpec("row", "integer", "Row number before which to insert (1-indexed)."),
                        ParamSpec("count", "integer", "Number of rows to insert.", required=False, default=1),
                    ],
                    returns="object",
                    returns_description='{"inserted": true, "row": int, "count": int}',
                    permission_required="local_worker",
                    tags=["row", "insert"],
                ),
                ActionSpec(
                    name="delete_rows",
                    description="Delete rows from a worksheet.",
                    params=[
                        ParamSpec("path", "string", "Path to the workbook."),
                        ParamSpec("sheet", "string", "Worksheet name."),
                        ParamSpec("row", "integer", "First row to delete (1-indexed)."),
                        ParamSpec("count", "integer", "Number of rows to delete.", required=False, default=1),
                    ],
                    returns="object",
                    returns_description='{"deleted": true, "row": int, "count": int}',
                    permission_required="local_worker",
                    tags=["row", "delete"],
                ),
                ActionSpec(
                    name="insert_columns",
                    description="Insert empty columns before a given column.",
                    params=[
                        ParamSpec("path", "string", "Path to the workbook."),
                        ParamSpec("sheet", "string", "Worksheet name."),
                        ParamSpec("column", "integer", "Column index before which to insert (1-indexed)."),
                        ParamSpec("count", "integer", "Number of columns to insert.", required=False, default=1),
                    ],
                    returns="object",
                    returns_description='{"inserted": true, "column": int, "count": int}',
                    permission_required="local_worker",
                    tags=["column", "insert"],
                ),
                ActionSpec(
                    name="delete_columns",
                    description="Delete columns from a worksheet.",
                    params=[
                        ParamSpec("path", "string", "Path to the workbook."),
                        ParamSpec("sheet", "string", "Worksheet name."),
                        ParamSpec("column", "integer", "First column to delete (1-indexed)."),
                        ParamSpec("count", "integer", "Number of columns to delete.", required=False, default=1),
                    ],
                    returns="object",
                    returns_description='{"deleted": true, "column": int, "count": int}',
                    permission_required="local_worker",
                    tags=["column", "delete"],
                ),
                ActionSpec(
                    name="set_column_width",
                    description="Set column width or auto-fit.",
                    params=[
                        ParamSpec("path", "string", "Path to the workbook."),
                        ParamSpec("sheet", "string", "Worksheet name."),
                        ParamSpec("column", "string", "Column letter or range, e.g. 'A' or 'A:D'."),
                        ParamSpec("width", "number", "Width in characters.", required=False),
                        ParamSpec("auto_fit", "boolean", "Auto-fit based on cell content.", required=False, default=False),
                    ],
                    returns="object",
                    returns_description='{"set": true, "column": str}',
                    permission_required="local_worker",
                    tags=["column", "width"],
                ),
                ActionSpec(
                    name="set_row_height",
                    description="Set the height of a row.",
                    params=[
                        ParamSpec("path", "string", "Path to the workbook."),
                        ParamSpec("sheet", "string", "Worksheet name."),
                        ParamSpec("row", "integer", "Row number (1-indexed)."),
                        ParamSpec("height", "number", "Row height in points."),
                    ],
                    returns="object",
                    returns_description='{"set": true, "row": int, "height": float}',
                    permission_required="local_worker",
                    tags=["row", "height"],
                ),
                # ── Cell manipulation ─────────────────────────────
                ActionSpec(
                    name="merge_cells",
                    description="Merge a range of cells into one.",
                    params=[
                        ParamSpec("path", "string", "Path to the workbook."),
                        ParamSpec("sheet", "string", "Worksheet name."),
                        ParamSpec("range", "string", "Range to merge, e.g. 'A1:C3'."),
                    ],
                    returns="object",
                    returns_description='{"merged": true, "range": str}',
                    permission_required="local_worker",
                    tags=["merge", "cell"],
                ),
                ActionSpec(
                    name="unmerge_cells",
                    description="Split a merged cell range back into individual cells.",
                    params=[
                        ParamSpec("path", "string", "Path to the workbook."),
                        ParamSpec("sheet", "string", "Worksheet name."),
                        ParamSpec("range", "string", "Merged range to split."),
                    ],
                    returns="object",
                    returns_description='{"unmerged": true, "range": str}',
                    permission_required="local_worker",
                    tags=["unmerge", "cell"],
                ),
                ActionSpec(
                    name="freeze_panes",
                    description="Freeze rows and/or columns at a given cell. Pass null to unfreeze.",
                    params=[
                        ParamSpec("path", "string", "Path to the workbook."),
                        ParamSpec("sheet", "string", "Worksheet name."),
                        ParamSpec("cell", "string", "Cell at the top-left of the frozen region, e.g. 'B2'. Null to unfreeze.", required=False),
                    ],
                    returns="object",
                    returns_description='{"frozen": true, "cell": str | null}',
                    permission_required="local_worker",
                    tags=["freeze", "panes"],
                ),
                # ── Data operations ───────────────────────────────
                ActionSpec(
                    name="find_replace",
                    description="Find and replace text across one or all worksheets.",
                    params=[
                        ParamSpec("path", "string", "Path to the workbook."),
                        ParamSpec("sheet", "string", "Worksheet name. All sheets if not set.", required=False),
                        ParamSpec("find", "string", "Text to search for."),
                        ParamSpec("replace", "string", "Replacement text."),
                        ParamSpec("case_sensitive", "boolean", "Case-sensitive search.", required=False, default=False),
                        ParamSpec("whole_cell", "boolean", "Match entire cell content only.", required=False, default=False),
                    ],
                    returns="object",
                    returns_description='{"replacements": int, "cells_checked": int}',
                    permission_required="local_worker",
                    tags=["find", "replace"],
                ),
                ActionSpec(
                    name="remove_duplicates",
                    description="Remove duplicate rows from a range.",
                    params=[
                        ParamSpec("path", "string", "Path to the workbook."),
                        ParamSpec("sheet", "string", "Worksheet name."),
                        ParamSpec("range", "string", "Range to check, e.g. 'A1:E100'."),
                        ParamSpec("columns", "array", "Column indices (0-based) to check. All if not set.", required=False),
                        ParamSpec("keep", "string", "Which duplicate to keep: 'first' or 'last'.", required=False, default="first"),
                    ],
                    returns="object",
                    returns_description='{"removed": int, "remaining": int}',
                    permission_required="local_worker",
                    tags=["duplicates", "data"],
                ),
                ActionSpec(
                    name="apply_formula",
                    description="Write an Excel formula to a cell.",
                    params=[
                        ParamSpec("path", "string", "Path to the workbook."),
                        ParamSpec("sheet", "string", "Worksheet name."),
                        ParamSpec("cell", "string", "Cell address.", example="C10"),
                        ParamSpec("formula", "string", "Excel formula, e.g. '=SUM(A1:A10)'.", example="=SUM(A1:A10)"),
                    ],
                    returns="object",
                    returns_description='{"applied": true, "cell": str, "formula": str}',
                    permission_required="local_worker",
                    tags=["formula"],
                ),
                ActionSpec(
                    name="add_named_range",
                    description="Define a named range in the workbook.",
                    params=[
                        ParamSpec("path", "string", "Path to the workbook."),
                        ParamSpec("name", "string", "Name for the range."),
                        ParamSpec("sheet", "string", "Worksheet name."),
                        ParamSpec("range", "string", "Cell range, e.g. 'A1:D100'."),
                        ParamSpec("scope", "string", "Scope: 'workbook' or 'sheet'.", required=False, default="workbook"),
                    ],
                    returns="object",
                    returns_description='{"added": true, "name": str}',
                    permission_required="local_worker",
                    tags=["named_range"],
                ),
                ActionSpec(
                    name="add_data_validation",
                    description="Add data validation rules to a range (dropdown list, number range, etc.).",
                    params=[
                        ParamSpec("path", "string", "Path to the workbook."),
                        ParamSpec("sheet", "string", "Worksheet name."),
                        ParamSpec("range", "string", "Cell range to validate.", example="B2:B100"),
                        ParamSpec("validation_type", "string", "Type: list, whole, decimal, date, time, textLength, custom.", required=False, default="list"),
                        ParamSpec("operator", "string", "Comparison operator.", required=False),
                        ParamSpec("formula1", "string", "First value / formula / list source.", required=False),
                        ParamSpec("formula2", "string", "Second value for 'between' ranges.", required=False),
                        ParamSpec("allow_blank", "boolean", "Allow blank cells.", required=False, default=True),
                        ParamSpec("show_dropdown", "boolean", "Show dropdown list.", required=False, default=True),
                        ParamSpec("show_error_alert", "boolean", "Show error alert on invalid input.", required=False, default=True),
                        ParamSpec("error_title", "string", "Error alert title.", required=False),
                        ParamSpec("error_message", "string", "Error alert message.", required=False),
                    ],
                    returns="object",
                    returns_description='{"added": true, "range": str, "type": str}',
                    permission_required="local_worker",
                    tags=["validation", "data"],
                ),
                ActionSpec(
                    name="add_autofilter",
                    description="Enable auto-filter (dropdown arrows) on a range.",
                    params=[
                        ParamSpec("path", "string", "Path to the workbook."),
                        ParamSpec("sheet", "string", "Worksheet name."),
                        ParamSpec("range", "string", "Range to apply filter. Uses used range if not set.", required=False),
                    ],
                    returns="object",
                    returns_description='{"added": true, "range": str}',
                    permission_required="local_worker",
                    tags=["filter", "data"],
                ),
                # ── Formatting ────────────────────────────────────
                ActionSpec(
                    name="format_range",
                    description=(
                        "Apply font, fill, border, and alignment formatting to a cell range. "
                        "Only specified properties are changed."
                    ),
                    params=[
                        ParamSpec("path", "string", "Path to the workbook."),
                        ParamSpec("sheet", "string", "Worksheet name."),
                        ParamSpec("range", "string", "Cell range to format.", example="A1:F1"),
                        ParamSpec("bold", "boolean", "Bold font.", required=False),
                        ParamSpec("italic", "boolean", "Italic font.", required=False),
                        ParamSpec("font_name", "string", "Font name, e.g. 'Arial'.", required=False),
                        ParamSpec("font_size", "integer", "Font size in points.", required=False),
                        ParamSpec("font_color", "string", "Hex font colour (e.g. 'FF0000').", required=False),
                        ParamSpec("fill_color", "string", "Hex background colour.", required=False),
                        ParamSpec("number_format", "string", "Number format string, e.g. '#,##0.00'.", required=False),
                        ParamSpec("alignment_horizontal", "string", "Horizontal alignment: left, center, right.", required=False),
                        ParamSpec("alignment_vertical", "string", "Vertical alignment: top, center, bottom.", required=False),
                        ParamSpec("wrap_text", "boolean", "Wrap text in cells.", required=False),
                        ParamSpec("border_style", "string", "Border style (thin, medium, thick, dashed, dotted, double, none).", required=False),
                        ParamSpec("border_color", "string", "Hex border colour.", required=False),
                    ],
                    returns="object",
                    returns_description='{"formatted": true, "cells_formatted": int}',
                    permission_required="local_worker",
                    tags=["format", "style"],
                ),
                ActionSpec(
                    name="apply_conditional_format",
                    description="Apply conditional formatting (color scales, data bars, icon sets, cell rules) to a range.",
                    params=[
                        ParamSpec("path", "string", "Path to the workbook."),
                        ParamSpec("sheet", "string", "Worksheet name."),
                        ParamSpec("range", "string", "Cell range.", example="A1:A100"),
                        ParamSpec("format_type", "string", "Type: color_scale, data_bar, icon_set, cell_is, formula.", required=False, default="cell_is"),
                        ParamSpec("operator", "string", "Comparison operator for cell_is.", required=False),
                        ParamSpec("value", "string", "Comparison value.", required=False),
                        ParamSpec("bold", "boolean", "Bold font in matched cells.", required=False),
                        ParamSpec("font_color", "string", "Font colour for matched cells.", required=False),
                        ParamSpec("fill_color", "string", "Fill colour for matched cells.", required=False),
                        ParamSpec("min_color", "string", "Color-scale minimum colour.", required=False, default="FFFFFF"),
                        ParamSpec("max_color", "string", "Color-scale maximum colour.", required=False, default="FF0000"),
                    ],
                    returns="object",
                    returns_description='{"applied": true, "range": str}',
                    permission_required="local_worker",
                    tags=["conditional", "format"],
                ),
                # ── Charts ────────────────────────────────────────
                ActionSpec(
                    name="create_chart",
                    description="Create a chart (bar, line, pie, doughnut, scatter, area, radar) from a data range and embed it in the sheet.",
                    params=[
                        ParamSpec("path", "string", "Path to the workbook."),
                        ParamSpec("sheet", "string", "Worksheet name."),
                        ParamSpec("chart_type", "string", "Chart type: bar, col, line, pie, doughnut, scatter, area, radar.", example="bar"),
                        ParamSpec("data_range", "string", "Data range for the chart.", example="A1:B10"),
                        ParamSpec("title", "string", "Chart title.", required=False),
                        ParamSpec("x_axis_title", "string", "X-axis title.", required=False),
                        ParamSpec("y_axis_title", "string", "Y-axis title.", required=False),
                        ParamSpec("position", "string", "Anchor cell for the chart.", required=False, default="E5"),
                        ParamSpec("width", "integer", "Chart width in centimetres.", required=False, default=15),
                        ParamSpec("height", "integer", "Chart height in centimetres.", required=False, default=10),
                        ParamSpec("has_legend", "boolean", "Show legend.", required=False, default=True),
                    ],
                    returns="object",
                    returns_description='{"created": true, "chart_type": str, "position": str}',
                    permission_required="local_worker",
                    tags=["chart", "visualisation"],
                ),
                # ── Images ────────────────────────────────────────
                ActionSpec(
                    name="insert_image",
                    description="Insert an image (PNG, JPEG, BMP, GIF) into a worksheet.",
                    params=[
                        ParamSpec("path", "string", "Path to the workbook."),
                        ParamSpec("sheet", "string", "Worksheet name."),
                        ParamSpec("image_path", "string", "Path to the image file."),
                        ParamSpec("cell", "string", "Cell to anchor the image.", example="B5"),
                        ParamSpec("width", "number", "Width in pixels.", required=False),
                        ParamSpec("height", "number", "Height in pixels.", required=False),
                    ],
                    returns="object",
                    returns_description='{"inserted": true, "cell": str}',
                    permission_required="local_worker",
                    tags=["image", "insert"],
                ),
                # ── Comments ──────────────────────────────────────
                ActionSpec(
                    name="add_comment",
                    description="Add a comment to a cell.",
                    params=[
                        ParamSpec("path", "string", "Path to the workbook."),
                        ParamSpec("sheet", "string", "Worksheet name."),
                        ParamSpec("cell", "string", "Cell address.", example="A1"),
                        ParamSpec("text", "string", "Comment text."),
                        ParamSpec("author", "string", "Comment author.", required=False, default="LLMOS Bridge"),
                    ],
                    returns="object",
                    returns_description='{"added": true, "cell": str}',
                    permission_required="local_worker",
                    tags=["comment", "add"],
                ),
                ActionSpec(
                    name="delete_comment",
                    description="Delete a comment from a cell.",
                    params=[
                        ParamSpec("path", "string", "Path to the workbook."),
                        ParamSpec("sheet", "string", "Worksheet name."),
                        ParamSpec("cell", "string", "Cell address."),
                    ],
                    returns="object",
                    returns_description='{"deleted": true, "cell": str}',
                    permission_required="local_worker",
                    tags=["comment", "delete"],
                ),
                # ── Export ────────────────────────────────────────
                ActionSpec(
                    name="export_to_csv",
                    description="Export a single worksheet to a CSV file.",
                    params=[
                        ParamSpec("path", "string", "Path to the workbook."),
                        ParamSpec("sheet", "string", "Worksheet name."),
                        ParamSpec("output_path", "string", "Destination CSV file path."),
                        ParamSpec("delimiter", "string", "CSV delimiter.", required=False, default=","),
                        ParamSpec("encoding", "string", "Output file encoding.", required=False, default="utf-8"),
                        ParamSpec("include_header", "boolean", "Include header row.", required=False, default=True),
                    ],
                    returns="object",
                    returns_description='{"exported": true, "output_path": str, "rows_written": int}',
                    permission_required="local_worker",
                    tags=["export", "csv"],
                ),
                ActionSpec(
                    name="export_to_pdf",
                    description="Convert the workbook to PDF using LibreOffice (must be installed).",
                    params=[
                        ParamSpec("path", "string", "Path to the workbook."),
                        ParamSpec("output_path", "string", "Destination PDF file path."),
                        ParamSpec("sheet", "string", "Sheet to export. All sheets if not set.", required=False),
                        ParamSpec("use_libreoffice", "boolean", "Use LibreOffice for conversion.", required=False, default=True),
                    ],
                    returns="object",
                    returns_description='{"exported": true, "output_path": str}',
                    permission_required="local_worker",
                    tags=["export", "pdf"],
                ),
            ],
        )
