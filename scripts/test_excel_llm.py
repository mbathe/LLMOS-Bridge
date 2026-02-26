#!/usr/bin/env python3
"""Test LLMOS Bridge Excel module with a real Claude LLM.

Creates a real .xlsx file on the desktop using ONLY native Excel actions.
The script tracks which tools Claude calls and flags any use of run_command
(which would mean Claude is bypassing native Excel tools).

Usage:
    ANTHROPIC_API_KEY=sk-... python scripts/test_excel_llm.py
"""

from __future__ import annotations

import json
import os
import sys
import tempfile
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent.parent / "packages" / "llmos-bridge"))
sys.path.insert(0, str(Path(__file__).parent.parent / "packages" / "langchain-llmos"))

OUTPUT_FILE = Path.home() / "Bureau" / "LLMOS_Budget_Test.xlsx"


def main() -> None:
    api_key = os.environ.get("ANTHROPIC_API_KEY")
    if not api_key:
        print("ERROR: Set ANTHROPIC_API_KEY environment variable")
        sys.exit(1)

    print("=" * 70)
    print("LLMOS Bridge — Excel Native Actions LLM Test")
    print("=" * 70)
    print(f"  Output: {OUTPUT_FILE}")
    print()

    from fastapi.testclient import TestClient

    from llmos_bridge.api.server import create_app
    from llmos_bridge.config import Settings

    tmp_dir = Path(tempfile.mkdtemp(prefix="llmos_excel_"))
    settings = Settings(
        memory={
            "state_db_path": str(tmp_dir / "state.db"),
            "vector_db_path": str(tmp_dir / "vector"),
        },
        logging={"level": "warning", "format": "console", "audit_file": None},
        modules={"enabled": ["excel"]},
        security={"require_approval_for": []},
    )
    app = create_app(settings=settings)

    with TestClient(app, raise_server_exceptions=True) as tc:
        # -----------------------------------------------------------------
        # Wire SDK
        # -----------------------------------------------------------------
        from langchain_llmos.client import LLMOSClient
        from langchain_llmos.toolkit import LLMOSToolkit

        toolkit = LLMOSToolkit.__new__(LLMOSToolkit)
        toolkit._base_url = str(tc.base_url)
        toolkit._api_token = None
        toolkit._timeout = 30.0
        toolkit._manifests = None
        toolkit._system_prompt = None
        toolkit._async_client = None

        client = LLMOSClient.__new__(LLMOSClient)
        client._http = tc
        client._base_url = str(tc.base_url)
        client._api_token = None
        client._timeout = 30.0
        toolkit._client = client

        tools = toolkit.get_tools(max_permission="local_worker")
        system_prompt = toolkit.get_system_prompt()

        excel_tools = [t for t in tools if t.module_id == "excel"]
        print(f"[1/3] Daemon started — {len(excel_tools)} Excel tools available")
        print()

        # -----------------------------------------------------------------
        # Build the prompt — a complex multi-step Excel task
        # -----------------------------------------------------------------
        from langchain_anthropic import ChatAnthropic
        from langchain_core.messages import HumanMessage, SystemMessage

        llm = ChatAnthropic(
            model="claude-sonnet-4-20250514",
            api_key=api_key,
            max_tokens=4096,
        )
        llm_with_tools = llm.bind_tools(tools)

        task = f"""Create a professional budget spreadsheet at: {OUTPUT_FILE}

Follow these steps EXACTLY using the native Excel tools available to you.
Do NOT use run_command or any workaround — use only the excel__* tools.

Step 1: Create a new workbook at the path above.

Step 2: Rename the default sheet to "Budget 2026".

Step 3: Write these headers in row 1 (A1:E1):
  Catégorie | Jan | Fév | Mar | Total

Step 4: Write this budget data starting at A2:
  Salaires     | 15000 | 15000 | 15500
  Loyer        | 3500  | 3500  | 3500
  Marketing    | 2000  | 2500  | 3000
  IT           | 1800  | 1800  | 2000
  Fournitures  | 500   | 600   | 450
  Transport    | 800   | 750   | 900

Step 5: Add SUM formulas in column E (E2:E7) for each row to total Jan+Fév+Mar.

Step 6: Add SUM formulas in row 9 (B9:E9) for column totals.
  Write "TOTAL" in A9.

Step 7: Format the header row (A1:E1): bold, blue background (4472C4), white font (FFFFFF).

Step 8: Format the total row (A9:E9): bold, light gray background (D9E2F3).

Step 9: Format columns B through E with number format "#,##0 €" (currency).

Step 10: Set column A width to 15 characters, and columns B-E to 12 characters.

Step 11: Freeze panes at B2 so headers and category names stay visible.

Step 12: Add a second sheet called "Résumé".

Step 13: Save the workbook.

IMPORTANT: Execute ALL steps. Use only the native excel__* tools."""

        print("[2/3] Sending task to Claude...")
        print("-" * 70)

        messages = [
            SystemMessage(content=system_prompt),
            HumanMessage(content=task),
        ]

        # -----------------------------------------------------------------
        # Agentic loop
        # -----------------------------------------------------------------
        tool_calls_log: list[dict] = []
        MAX_TURNS = 25
        bypasses = []

        for turn in range(MAX_TURNS):
            response = llm_with_tools.invoke(messages)
            messages.append(response)

            if not response.tool_calls:
                print(f"\n  [Turn {turn + 1}] Claude finished:")
                print(f"  {response.content[:500]}")
                break

            print(f"  [Turn {turn + 1}] {len(response.tool_calls)} tool call(s)")
            for tc_item in response.tool_calls:
                name = tc_item["name"]
                args = tc_item["args"]
                args_str = json.dumps(args, ensure_ascii=False)
                if len(args_str) > 120:
                    args_str = args_str[:117] + "..."

                # Track if Claude bypasses native tools
                is_bypass = "run_command" in name or "os_exec" in name
                marker = " *** BYPASS ***" if is_bypass else ""
                if is_bypass:
                    bypasses.append(name)

                print(f"    {name}({args_str}){marker}")
                tool_calls_log.append({"name": name, "args": tc_item["args"]})

                tool = next((t for t in tools if t.name == name), None)
                if tool:
                    try:
                        result = tool.invoke(args)
                        result_str = result[:150] if len(result) > 150 else result
                        print(f"      → {result_str}")
                    except Exception as exc:
                        result = json.dumps({"error": str(exc)})
                        print(f"      → ERROR: {exc}")
                    from langchain_core.messages import ToolMessage

                    messages.append(
                        ToolMessage(content=result, tool_call_id=tc_item["id"])
                    )
        else:
            print(f"\n  (Stopped after {MAX_TURNS} turns)")

        # -----------------------------------------------------------------
        # Verification
        # -----------------------------------------------------------------
        print()
        print("=" * 70)
        print("[3/3] VERIFICATION")
        print("=" * 70)

        # 1. File exists
        file_exists = OUTPUT_FILE.exists()
        print(f"  File exists: {'YES' if file_exists else 'NO'}")

        if file_exists:
            import openpyxl

            wb = openpyxl.load_workbook(str(OUTPUT_FILE), data_only=False)
            print(f"  Sheets: {wb.sheetnames}")

            # Check "Budget 2026" sheet
            has_budget = "Budget 2026" in wb.sheetnames
            print(f"  Has 'Budget 2026' sheet: {'YES' if has_budget else 'NO'}")

            has_resume = "Résumé" in wb.sheetnames
            print(f"  Has 'Résumé' sheet: {'YES' if has_resume else 'NO'}")

            if has_budget:
                ws = wb["Budget 2026"]
                # Check headers
                headers = [ws.cell(1, c).value for c in range(1, 6)]
                print(f"  Headers: {headers}")

                # Check data (A2:D7)
                a2 = ws["A2"].value
                b2 = ws["B2"].value
                print(f"  A2 (first category): {a2}")
                print(f"  B2 (first value): {b2}")

                # Check formulas in E column
                e2 = ws["E2"].value
                print(f"  E2 (formula): {e2}")

                # Check total row
                a9 = ws["A9"].value
                b9 = ws["B9"].value
                print(f"  A9: {a9}")
                print(f"  B9 (total formula): {b9}")

                # Check formatting on header
                a1_font = ws["A1"].font
                a1_fill = ws["A1"].fill
                print(f"  A1 font bold: {a1_font.bold}")
                print(f"  A1 fill color: {a1_fill.fgColor.rgb if a1_fill.fgColor else 'none'}")

                # Check freeze panes
                print(f"  Freeze panes: {ws.freeze_panes}")

                # Row count with data
                data_rows = sum(1 for row in ws.iter_rows(min_row=2, max_row=8, max_col=1) if row[0].value)
                print(f"  Data rows (A2:A8): {data_rows}")

            wb.close()

        # 2. Tool usage analysis
        print()
        total_calls = len(tool_calls_log)
        excel_calls = sum(1 for c in tool_calls_log if c["name"].startswith("excel__"))
        bypass_calls = len(bypasses)

        print(f"  Total tool calls: {total_calls}")
        print(f"  Excel native calls: {excel_calls}")
        print(f"  Bypass calls (run_command/os_exec): {bypass_calls}")

        # Unique actions used
        unique_actions = sorted(set(c["name"] for c in tool_calls_log))
        print(f"  Unique tools used ({len(unique_actions)}):")
        for a in unique_actions:
            count = sum(1 for c in tool_calls_log if c["name"] == a)
            print(f"    {a}: {count}x")

        # Final verdict
        print()
        checks = {
            "File created": file_exists,
            "Budget 2026 sheet": file_exists and has_budget,
            "Résumé sheet": file_exists and has_resume,
            "No bypasses": bypass_calls == 0,
            "Used native Excel tools": excel_calls > 0,
        }

        all_pass = all(checks.values())
        for check, passed in checks.items():
            status = "PASS" if passed else "FAIL"
            print(f"  [{status}] {check}")

        print()
        if all_pass:
            print("  RESULT: ALL CHECKS PASSED")
            print(f"  Open the file to verify: {OUTPUT_FILE}")
        else:
            print("  RESULT: SOME CHECKS FAILED")

        print("=" * 70)


if __name__ == "__main__":
    main()
